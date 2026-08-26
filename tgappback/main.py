import os
import uuid
import logging
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Optional

import socketio
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import select, delete, func, text, update
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from slowapi.errors import RateLimitExceeded
from fastapi.responses import JSONResponse

from models import async_session, Product, Category, ProductImage, Order, OrderItem, init_db, User, UserIdentity, AuthProvider, UserRole, Courier, OrderStatus, ProductPhotoGroup, Promo, Margin, Review, Faq, SupportConversation, SupportMessage
from auth import hash_password, verify_password, router as auth_router, get_current_user, get_optional_user, decode_token
from rate_limit import limiter
from parserMAIN import parse_iphone_excel_to_dicts, parse_macbook_excel_to_dicts, parse_ipad_excel_to_dicts, parse_imac_excel_to_dicts, parse_watch_excel_to_dicts, parse_airpods_excel_to_dicts, parse_accessories_excel_to_dicts, parse_samsung_excel_to_dicts, parse_consoles_excel_to_dicts, parse_dyson_excel_to_dicts, parse_xiaomi_excel_to_dicts, parse_poco_excel_to_dicts, parse_stations_excel_to_dicts
import openpyxl
import httpx

logger = logging.getLogger("uvicorn.error")

# Допустимые origins (dev + продакшн домены из env)
ALLOWED_ORIGINS = [
    o.strip()
    for o in os.getenv("ALLOWED_ORIGINS", "http://localhost:5173,http://127.0.0.1:5173").split(",")
    if o.strip()
]

fastapi_app = FastAPI()

# === Socket.IO ===
sio = socketio.AsyncServer(async_mode="asgi", cors_allowed_origins=ALLOWED_ORIGINS, ping_timeout=60, ping_interval=25)

@sio.event
async def connect(sid, environ, auth):
    token = auth.get("token") if auth else None
    if not token:
        raise socketio.exceptions.ConnectionRefusedError("No token")
    try:
        user_id = decode_token(token)
    except Exception:
        raise socketio.exceptions.ConnectionRefusedError("Invalid token")

    async with async_session() as session:
        if isinstance(user_id, dict) and user_id.get("admin"):
            role = "admin"
        else:
            result = await session.execute(
                select(User).options(selectinload(User.auth_identities)).where(User.id == user_id)
            )
            user = result.scalar_one_or_none()
            if not user:
                raise socketio.exceptions.ConnectionRefusedError("User not found")
            role = user.role.value

        async with sio.session(sid) as session_data:
            session_data["user_id"] = user_id
            session_data["role"] = role

    await sio.enter_room(sid, f"{role}_room")
    if role == "admin":
        await sio.enter_room(sid, "admin_room")
    elif role == "courier":
        await sio.enter_room(sid, "courier_room")
    elif isinstance(user_id, int):
        await sio.enter_room(sid, f"user_{user_id}")

@sio.event
async def order_status_changed(sid, data):
    async with sio.session(sid) as session_data:
        role = session_data.get("role")
        user_id = session_data.get("user_id")

    if role not in ("admin", "courier"):
        return

    order_id = data.get("order_id")
    new_status = data.get("status")
    if not order_id or not new_status:
        return
    try:
        new_status = OrderStatus(new_status)
    except ValueError:
        return

    async with async_session() as db:
        result = await db.execute(select(Order).where(Order.id == order_id))
        order = result.scalar_one_or_none()
        if not order:
            return
        if role == "courier":
            courier = await _get_courier_for_user(db, user_id)
            if not courier or order.courier_id != courier.id:
                return
        order.status = new_status
        order.status_changed_at = datetime.utcnow()
        await db.commit()
        await _notify_order_status(order, new_status.value)

    # Broadcast to admin room
    await sio.emit("order_updated", {
        "order_id": order_id,
        "status": new_status.value,
    }, room="admin_room")

@sio.event
async def disconnect(sid):
    pass

# Mount FastAPI as sub-app under socket.io
app = socketio.ASGIApp(sio, other_asgi_app=fastapi_app)

# Make the fastapi_app accessible for middleware/routes below
fastapi_app.state.limiter = limiter
fastapi_app.add_exception_handler(
    RateLimitExceeded,
    lambda req, exc: JSONResponse(
        status_code=429,
        content={"detail": "Слишком много попыток. Попробуйте позже."},
    ),
)
fastapi_app.include_router(auth_router)

fastapi_app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)
fastapi_app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")

ALLOWED_UPLOAD_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
MAX_UPLOAD_SIZE = 10 * 1024 * 1024


@fastapi_app.get("/api/health")
async def health():
    return {"status": "ok"}


def _is_allowed_image(content: bytes) -> bool:
    """Проверка сигнатуры файла (magic bytes): jpg/png/webp/gif."""
    if content[:3] == b"\xff\xd8\xff":
        return True
    if content[:8] == b"\x89PNG\r\n\x1a\n":
        return True
    if content[:4] == b"RIFF" and content[8:12] == b"WEBP":
        return True
    if content[:6] in (b"GIF87a", b"GIF89a"):
        return True
    return False


async def get_db():
    async with async_session() as session:
        yield session


async def require_admin(user: User = Depends(get_current_user)) -> User:
    if user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")
    return user


class ProductCreate(BaseModel):
    name: str
    price: Optional[float] = None
    purchase_price: Optional[float] = None
    is_available: bool = True
    quantity: int = 1
    category_id: int
    attributes: dict = {}
    images: list[dict] = []


class ProductUpdate(BaseModel):
    name: Optional[str] = None
    price: Optional[float] = None
    purchase_price: Optional[float] = None
    is_available: Optional[bool] = None
    quantity: Optional[int] = None
    category_id: Optional[int] = None
    attributes: Optional[dict] = None
    images: Optional[list[dict]] = None


class SimilarRequest(BaseModel):
    name: str
    color: str
    exclude_id: int


class CategoryCreate(BaseModel):
    name: str


@fastapi_app.on_event("startup")
async def on_startup():
    await init_db()
    try:
        async with async_session() as session:
            await _verify_upload_files(session)
    except Exception as e:
        logger.warning("Не удалось проверить целостность картинок: %s", e)
    # Seed example FAQ items (only if table is empty)
    try:
        async with async_session() as session:
            result = await session.execute(select(func.count()).select_from(Faq))
            if result.scalar_one() == 0:
                session.add_all([
                    Faq(
                        question="Какие сроки доставки?",
                        answer="Доставка по Москве занимает 1–2 дня, по России — от 3 до 7 дней в зависимости от региона. Точный срок менеджер сообщит после оформления заказа.",
                    ),
                    Faq(
                        question="Как я могу оплатить заказ?",
                        answer="Мы принимаем наличные при получении, перевод на карту и оплату онлайн. Способ оплаты вы выбираете при оформлении заказа.",
                    ),
                    Faq(
                        question="Можно ли вернуть товар?",
                        answer="Да, в течение 14 дней с момента покупки при сохранении товарного вида и упаковки. Для оформления возврата свяжитесь с нашей поддержкой.",
                    ),
                ])
                await session.commit()
    except Exception as e:
        logger.warning("Не удалось заполнить FAQ: %s", e)
    # Add GOOGLE to authprovider enum if not present (migration for existing DB)
    try:
        async with async_session() as session:
            await session.execute(text("ALTER TYPE authprovider ADD VALUE IF NOT EXISTS 'GOOGLE'"))
            await session.commit()
    except Exception:
        pass
    # Migration: create couriers table if not exists and add courier_id to orders
    try:
        async with async_session() as session:
            await session.execute(text("""
                CREATE TABLE IF NOT EXISTS couriers (
                    id SERIAL PRIMARY KEY,
                    name VARCHAR(200) NOT NULL,
                    phone VARCHAR(20),
                    login VARCHAR(100) UNIQUE NOT NULL,
                    password_hash VARCHAR(255) NOT NULL,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS courier_id INTEGER REFERENCES couriers(id) ON DELETE SET NULL
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE support_messages ADD COLUMN IF NOT EXISTS images JSON
            """))
            await session.execute(text("""
                ALTER TABLE support_conversations ADD COLUMN IF NOT EXISTS user_last_read_at TIMESTAMP
            """))
            await session.execute(text("""
                ALTER TABLE support_conversations ADD COLUMN IF NOT EXISTS admin_last_read_at TIMESTAMP
            """))
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS status_changed_at TIMESTAMP
            """))
            await session.execute(text("""
                ALTER TABLE users ADD COLUMN IF NOT EXISTS last_orders_seen_at TIMESTAMP
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("ALTER TABLE faq ADD COLUMN IF NOT EXISTS position INTEGER"))
            await session.execute(text("UPDATE faq SET position = id WHERE position IS NULL"))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS confirmation_code VARCHAR(5)
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE orders DROP COLUMN IF EXISTS gift_name
            """))
            await session.execute(text("""
                ALTER TABLE orders DROP COLUMN IF EXISTS gift_image
            """))
            await session.execute(text("""
                ALTER TABLE orders DROP COLUMN IF EXISTS gift_price
            """))
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS gifts JSONB
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS delivery_photo_urls JSONB
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS delivery_imei VARCHAR(255)
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS delivered_at TIMESTAMP
            """))
            await session.commit()
    except Exception:
        pass
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE orders ADD COLUMN IF NOT EXISTS archived_at TIMESTAMP
            """))
            await session.commit()
    except Exception:
        pass
    # Миграция: цена закупки для товаров + цена продажи становится необязательной
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE products ADD COLUMN IF NOT EXISTS purchase_price DOUBLE PRECISION
            """))
            # Переносим текущую цену в цену закупки (раньше импорт Excel
            # записывал закупочную цену прямо в цену продажи).
            await session.execute(text("""
                UPDATE products SET purchase_price = price WHERE purchase_price IS NULL
            """))
            await session.execute(text("""
                ALTER TABLE products ALTER COLUMN price DROP NOT NULL
            """))
            await session.commit()
    except Exception:
        pass
    # Backfill фото-групп по (наименование, цвет) из уже загруженных товаров
    try:
        async with async_session() as session:
            await _ensure_photo_groups_from_products(session)
    except Exception:
        pass
    # Снапшоты позиций заказа + мягкое удаление товаров (SET NULL вместо RESTRICT)
    try:
        async with async_session() as session:
            await session.execute(text("""
                ALTER TABLE order_items ADD COLUMN IF NOT EXISTS product_name VARCHAR(255)
            """))
            await session.execute(text("""
                ALTER TABLE order_items ADD COLUMN IF NOT EXISTS product_image VARCHAR(1000)
            """))
            await session.execute(text("""
                ALTER TABLE order_items DROP CONSTRAINT IF EXISTS order_items_product_id_fkey
            """))
            await session.execute(text("""
                ALTER TABLE order_items ALTER COLUMN product_id DROP NOT NULL
            """))
            await session.execute(text("""
                ALTER TABLE order_items
                    ADD CONSTRAINT order_items_product_id_fkey
                    FOREIGN KEY (product_id) REFERENCES products(id) ON DELETE SET NULL
            """))
            await session.execute(text("""
                UPDATE order_items oi SET
                    product_name = COALESCE(oi.product_name, p.name),
                    product_image = COALESCE(oi.product_image, (
                        SELECT pi.image_url FROM product_images pi
                        WHERE pi.product_id = p.id ORDER BY pi.is_main DESC, pi.id LIMIT 1
                    ))
                FROM products p WHERE oi.product_id = p.id
            """))
            await session.commit()
    except Exception:
        pass
    await seed_demo_users()
    await remove_admin_from_db()
    await migrate_orders_user_id()


async def remove_admin_from_db():
    """Remove any admin users from DB — admin is only in .env."""
    async with async_session() as session:
        result = await session.execute(
            select(User).where(User.role == UserRole.ADMIN)
        )
        admins = list(result.scalars().all())
        if not admins:
            return
        for admin in admins:
            await session.delete(admin)
        await session.commit()
        print(f"[CLEANUP] Removed {len(admins)} admin user(s) from DB (admin is now env-only)")


async def migrate_orders_user_id():
    """Assign user_id to existing orders with NULL user_id by matching customer_name against User.name or User.username."""
    async with async_session() as session:
        result = await session.execute(select(Order).where(Order.user_id == None))
        orphan_orders = list(result.scalars().all())
        if not orphan_orders:
            return

        users_result = await session.execute(select(User))
        users = list(users_result.scalars().all())

        updated = 0
        for order in orphan_orders:
            if not order.customer_name:
                continue
            name_lower = order.customer_name.lower().strip()
            for user in users:
                if (user.name and user.name.lower().strip() == name_lower) or \
                   (user.username and user.username.lower().strip() == name_lower):
                    order.user_id = user.id
                    updated += 1
                    break

        if updated:
            await session.commit()
            print(f"[MIGRATE] Assigned user_id to {updated} orders by customer_name")


async def seed_demo_users():
    """Создаёт/обновляет тестовый аккаунт из .env (TEST_USER_*), если он задан."""
    login = os.getenv("TEST_USER_LOGIN", "").strip()
    password = os.getenv("TEST_USER_PASSWORD", "")
    email = os.getenv("TEST_USER_EMAIL", "").strip() or None
    name = os.getenv("TEST_USER_NAME", "Тестовый пользователь").strip()
    phone = os.getenv("TEST_USER_PHONE", "").strip() or None
    if not login or not password:
        return

    async with async_session() as session:
        result = await session.execute(
            select(UserIdentity).where(
                UserIdentity.provider == AuthProvider.LOCAL,
                UserIdentity.provider_user_id.in_([v for v in (login, email) if v]),
            )
        )
        existing = result.scalar_one_or_none()
        if existing:
            if not verify_password(password, existing.password_hash):
                existing.password_hash = hash_password(password)
                await session.commit()
            return

        user = User(name=name, email=email, phone=phone, username=login, role=UserRole.USER)
        session.add(user)
        await session.flush()
        session.add(UserIdentity(
            user_id=user.id,
            provider=AuthProvider.LOCAL,
            provider_user_id=email or login,
            password_hash=hash_password(password),
        ))
        await session.commit()


# === Cart Merge ===
class CartMergeItem(BaseModel):
    id: int
    quantity: int

class CartMergeRequest(BaseModel):
    items: list[CartMergeItem]

@fastapi_app.post("/api/cart/merge")
async def cart_merge(req: CartMergeRequest, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    # Check all products exist
    for item in req.items:
        result = await db.execute(select(Product).where(Product.id == item.id))
        if not result.scalar_one_or_none():
            raise HTTPException(status_code=404, detail=f"Product {item.id} not found")
    # If user has an existing order with status "cart", update it; otherwise create one
    result = await db.execute(
        select(Order).where(Order.user_id == user.id, Order.status == "cart").options(selectinload(Order.items))
    )
    cart = result.scalar_one_or_none()
    if cart:
        cart.items.clear()
        for item in req.items:
            cart.items.append(OrderItem(product_id=item.id, quantity=item.quantity, price=0))
    else:
        cart = Order(user_id=user.id, status="cart", total=0, items=[])
        db.add(cart)
        await db.flush()
        for item in req.items:
            cart.items.append(OrderItem(product_id=item.id, quantity=item.quantity, price=0))
    await db.commit()
    return {"ok": True, "cart_id": cart.id}

# === Activity (Yandex Metrika) ===
class ActivityResponse(BaseModel):
    visitors_by_day: list[dict]
    sources: list[dict]
    devices: list[dict]
    summary: dict


YANDEX_METRIKA_COUNTER = os.getenv("YANDEX_METRIKA_COUNTER")
YANDEX_METRIKA_TOKEN = os.getenv("YANDEX_METRIKA_TOKEN")

METRIKA_BASE = "https://api-metrika.yandex.net/stat/v1/data"


async def fetch_metrika(metrics: str, dimensions: str, date1: str, date2: str) -> dict | None:
    if not YANDEX_METRIKA_COUNTER or not YANDEX_METRIKA_TOKEN:
        return None
    async with httpx.AsyncClient() as client:
        r = await client.get(
            METRIKA_BASE,
            params={
                "ids": YANDEX_METRIKA_COUNTER,
                "date1": date1,
                "date2": date2,
                "metrics": metrics,
                "dimensions": dimensions,
                "accuracy": "full",
            },
            headers={"Authorization": f"OAuth {YANDEX_METRIKA_TOKEN}"},
            timeout=15,
        )
        if r.status_code != 200:
            return None
        return r.json()


@fastapi_app.get("/api/admin/activity")
async def get_activity(
    days: int = 7,
    user: User = Depends(get_current_user),
):
    if user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    if not YANDEX_METRIKA_COUNTER or not YANDEX_METRIKA_TOKEN:
        return ActivityResponse(
            visitors_by_day=[],
            sources=[],
            devices=[],
            summary={"total_visitors": 0, "total_visits": 0, "total_pageviews": 0, "avg_duration": 0, "bounce_rate": 0},
        )

    today = datetime.utcnow().strftime("%Y-%m-%d")
    start = (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")

    # Visitors by day
    by_day = await fetch_metrika(
        "ym:s:visitors,ym:s:visits,ym:s:pageviews",
        "ym:s:date",
        start, today,
    )

    # Traffic sources
    sources = await fetch_metrika(
        "ym:s:visits,ym:s:visitors",
        "ym:s:trafficSource",
        start, today,
    )

    # Devices
    devices = await fetch_metrika(
        "ym:s:visits,ym:s:visitors",
        "ym:s:deviceCategory",
        start, today,
    )

    # Summary
    summary_data = await fetch_metrika(
        "ym:s:visitors,ym:s:visits,ym:s:pageviews,ym:s:avgVisitDurationSeconds,ym:s:bounceRate",
        "",
        start, today,
    )

    def extract_totals(data: dict | None, metrics: list[str]) -> list[float]:
        if not data or "totals" not in data:
            return [0.0] * len(metrics)
        return data["totals"]

    summary_totals = extract_totals(summary_data, ["visitors", "visits", "pageviews", "avg_duration", "bounce_rate"])

    def map_rows(data: dict | None, dim_key: str, metric_keys: list[str]) -> list[dict]:
        if not data or "data" not in data:
            return []
        rows = []
        for entry in data["data"]:
            dims = entry.get("dimensions", [])
            label = dims[0]["name"] if dims else "unknown"
            vals = entry.get("metrics", [])
            row = {dim_key: label}
            for i, key in enumerate(metric_keys):
                row[key] = vals[i] if i < len(vals) else 0
            rows.append(row)
        return rows

    return ActivityResponse(
        visitors_by_day=map_rows(by_day, "date", ["visitors", "visits", "pageviews"]),
        sources=map_rows(sources, "source", ["visits", "visitors"]),
        devices=map_rows(devices, "device", ["visits", "visitors"]),
        summary={
            "total_visitors": summary_totals[0] if len(summary_totals) > 0 else 0,
            "total_visits": summary_totals[1] if len(summary_totals) > 1 else 0,
            "total_pageviews": summary_totals[2] if len(summary_totals) > 2 else 0,
            "avg_duration": summary_totals[3] if len(summary_totals) > 3 else 0,
            "bounce_rate": summary_totals[4] if len(summary_totals) > 4 else 0,
        },
    )


### ОТЗЫВЫ


class ReviewCreate(BaseModel):
    rating: int
    text: str
    images: list[str] = []


class ReviewStatusUpdate(BaseModel):
    status: str


def _review_to_dict(r: Review) -> dict:
    return {
        "id": r.id,
        "user_id": r.user_id,
        "user_name": r.user_name,
        "rating": r.rating,
        "text": r.text,
        "images": r.images or [],
        "status": r.status,
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


@fastapi_app.get("/api/reviews")
async def list_reviews(db: AsyncSession = Depends(get_db)):
    """Публичный список опубликованных отзывов + общее количество и средняя оценка."""
    result = await db.execute(
        select(Review)
        .where(Review.status == "published")
        .order_by(Review.created_at.desc())
    )
    reviews = result.scalars().all()
    avg_rating = round(sum(r.rating for r in reviews) / len(reviews), 1) if reviews else 0
    return {
        "total": len(reviews),
        "avg_rating": avg_rating,
        "reviews": [_review_to_dict(r) for r in reviews],
    }


@fastapi_app.post("/api/reviews")
async def create_review(
    data: ReviewCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Создать отзыв — только для авторизованных пользователей."""
    if data.rating < 1 or data.rating > 5:
        raise HTTPException(400, "Оценка должна быть от 1 до 5")
    text = data.text.strip()
    if len(text) < 2:
        raise HTTPException(400, "Напишите текст отзыва")

    review = Review(
        user_id=user.id,
        user_name=user.name or "Пользователь",
        rating=data.rating,
        text=text,
        images=data.images or [],
        status="new",
    )
    db.add(review)
    await db.commit()
    return _review_to_dict(review)


@fastapi_app.get("/api/admin/reviews")
async def list_admin_reviews(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")
    result = await db.execute(select(Review).order_by(Review.created_at.desc()))
    return [_review_to_dict(r) for r in result.scalars().all()]


@fastapi_app.patch("/api/admin/reviews/{review_id}")
async def update_review_status(
    review_id: int,
    data: ReviewStatusUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")
    if data.status not in ("new", "published", "hidden"):
        raise HTTPException(400, "Недопустимый статус")

    result = await db.execute(select(Review).where(Review.id == review_id))
    review = result.scalar_one_or_none()
    if not review:
        raise HTTPException(404, "Отзыв не найден")

    review.status = data.status
    await db.commit()
    return _review_to_dict(review)


@fastapi_app.delete("/api/admin/reviews/{review_id}")
async def delete_review(
    review_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Review).where(Review.id == review_id))
    review = result.scalar_one_or_none()
    if not review:
        raise HTTPException(404, "Отзыв не найден")

    await db.delete(review)
    await db.commit()
    return {"ok": True}


@fastapi_app.delete("/api/admin/reviews")
async def delete_all_admin_reviews(
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")
    if status is not None and status not in ("new", "published", "hidden"):
        raise HTTPException(400, "Недопустимый статус")

    query = select(Review)
    if status is not None:
        query = query.where(Review.status == status)
    result = await db.execute(query)
    reviews = result.scalars().all()
    for review in reviews:
        await db.delete(review)
    await db.commit()
    return {"ok": True, "deleted": len(reviews)}


### FAQ (часто задаваемые вопросы)


class FaqCreate(BaseModel):
    question: str
    answer: str


class FaqReorder(BaseModel):
    ids: list[int]


def _faq_to_dict(f: Faq) -> dict:
    return {
        "id": f.id,
        "question": f.question,
        "answer": f.answer,
        "position": f.position or 0,
        "created_at": f.created_at.isoformat() if f.created_at else None,
    }


@fastapi_app.get("/api/faq")
async def list_faq(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Faq).order_by(Faq.position.asc(), Faq.id.asc()))
    return [_faq_to_dict(f) for f in result.scalars().all()]


@fastapi_app.put("/api/admin/faq/reorder")
async def reorder_faq(
    data: FaqReorder,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    for position, faq_id in enumerate(data.ids):
        result = await db.execute(select(Faq).where(Faq.id == faq_id))
        faq = result.scalar_one_or_none()
        if faq:
            faq.position = position
    await db.commit()
    return {"ok": True}


@fastapi_app.post("/api/admin/faq")
async def create_faq(
    data: FaqCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")
    question = data.question.strip()
    answer = data.answer.strip()
    if not question:
        raise HTTPException(400, "Вопрос не может быть пустым")
    if not answer:
        raise HTTPException(400, "Ответ не может быть пустым")

    max_pos_result = await db.execute(select(func.max(Faq.position)))
    max_pos = max_pos_result.scalar()

    faq = Faq(question=question, answer=answer, position=(max_pos or 0) + 1)
    db.add(faq)
    await db.commit()
    await db.refresh(faq)
    return _faq_to_dict(faq)


@fastapi_app.patch("/api/admin/faq/{faq_id}")
async def update_faq(
    faq_id: int,
    data: FaqCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Faq).where(Faq.id == faq_id))
    faq = result.scalar_one_or_none()
    if not faq:
        raise HTTPException(404, "Вопрос не найден")

    faq.question = data.question.strip()
    faq.answer = data.answer.strip()
    await db.commit()
    return _faq_to_dict(faq)


@fastapi_app.delete("/api/admin/faq/{faq_id}")
async def delete_faq(
    faq_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Faq).where(Faq.id == faq_id))
    faq = result.scalar_one_or_none()
    if not faq:
        raise HTTPException(404, "Вопрос не найден")

    await db.delete(faq)
    await db.commit()
    return {"ok": True}


@fastapi_app.delete("/api/admin/faq")
async def delete_all_faq(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Faq))
    faqs = result.scalars().all()
    for faq in faqs:
        await db.delete(faq)
    await db.commit()
    return {"ok": True, "deleted": len(faqs)}


### ПОДДЕРЖКА (диалоги пользователей)


class SupportMessageCreate(BaseModel):
    text: str
    images: list[str] = []


def _msg_to_dict(m: SupportMessage) -> dict:
    return {
        "id": m.id,
        "conversation_id": m.conversation_id,
        "sender": m.sender,
        "text": m.text,
        "images": m.images or [],
        "created_at": m.created_at.isoformat() if m.created_at else None,
    }


def _conv_to_dict(c: SupportConversation, with_messages: bool = False) -> dict:
    last = c.messages[-1] if c.messages else None
    unread = False
    admin_unread_count = 0
    if c.messages:
        unread = any(
            m.sender == "admin"
            and (c.user_last_read_at is None or m.created_at > c.user_last_read_at)
            for m in c.messages
        )
        admin_unread_count = sum(
            1
            for m in c.messages
            if m.sender == "user"
            and (c.admin_last_read_at is None or m.created_at > c.admin_last_read_at)
        )
    return {
        "id": c.id,
        "user_id": c.user_id,
        "user_name": c.user_name,
        "status": c.status,
        "unread": unread,
        "admin_unread_count": admin_unread_count,
        "created_at": c.created_at.isoformat() if c.created_at else None,
        "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        "last_message": last.text if last else None,
        "last_message_at": last.created_at.isoformat() if last and last.created_at else None,
        "messages": [_msg_to_dict(m) for m in c.messages] if with_messages else None,
    }


async def _load_conversation_with_messages(db: AsyncSession, conv_id: int) -> SupportConversation:
    result = await db.execute(
        select(SupportConversation)
        .options(selectinload(SupportConversation.messages))
        .where(SupportConversation.id == conv_id)
    )
    return result.scalar_one_or_none()


@fastapi_app.post("/api/support/conversations")
async def create_conversation(
    data: SupportMessageCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    text = data.text.strip()
    if not text:
        raise HTTPException(400, "Сообщение не может быть пустым")

    # У одного пользователя один диалог: переиспользуем его, если он уже есть
    result = await db.execute(
        select(SupportConversation)
        .where(SupportConversation.user_id == current_user.id)
        .order_by(SupportConversation.updated_at.desc())
    )
    conv = result.scalars().first()
    if conv is None:
        conv = SupportConversation(
            user_id=None if current_user.id == 0 else current_user.id,
            user_name=current_user.name or current_user.username or "Пользователь",
            status="open",
        )
        db.add(conv)
        await db.flush()

    msg = SupportMessage(conversation_id=conv.id, sender="user", text=text, images=data.images or [])
    db.add(msg)
    conv.user_last_read_at = datetime.utcnow()
    await db.commit()
    await db.refresh(msg)

    await sio.emit(
        "support_user_message",
        {"conversation_id": conv.id, "message": _msg_to_dict(msg)},
        room="admin_room",
    )

    loaded = await _load_conversation_with_messages(db, conv.id)
    return _conv_to_dict(loaded, with_messages=True)


@fastapi_app.get("/api/support/conversations")
async def list_my_conversations(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(SupportConversation)
        .options(selectinload(SupportConversation.messages))
        .where(SupportConversation.user_id == current_user.id)
        .order_by(SupportConversation.updated_at.desc())
    )
    return [_conv_to_dict(c, with_messages=True) for c in result.scalars().all()]


@fastapi_app.post("/api/support/conversations/{conv_id}/messages")
async def send_user_message(
    conv_id: int,
    data: SupportMessageCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    text = data.text.strip()
    if not text:
        raise HTTPException(400, "Сообщение не может быть пустым")

    conv = await _load_conversation_with_messages(db, conv_id)
    if not conv:
        raise HTTPException(404, "Диалог не найден")
    if conv.user_id != current_user.id:
        raise HTTPException(403, "Нет доступа к этому диалогу")

    msg = SupportMessage(conversation_id=conv.id, sender="user", text=text, images=data.images or [])
    db.add(msg)
    conv.user_last_read_at = datetime.utcnow()
    await db.commit()
    await db.refresh(msg)
    await sio.emit(
        "support_user_message",
        {"conversation_id": conv.id, "message": _msg_to_dict(msg)},
        room="admin_room",
    )
    return _msg_to_dict(msg)


@fastapi_app.post("/api/support/conversations/{conv_id}/read")
async def mark_conversation_read(
    conv_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    conv = await _load_conversation_with_messages(db, conv_id)
    if not conv:
        raise HTTPException(404, "Диалог не найден")
    if conv.user_id != current_user.id:
        raise HTTPException(403, "Нет доступа к этому диалогу")

    conv.user_last_read_at = datetime.utcnow()
    await db.commit()
    return {"ok": True}


@fastapi_app.get("/api/support/unread")
async def my_support_unread(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(SupportConversation)
        .options(selectinload(SupportConversation.messages))
        .where(SupportConversation.user_id == current_user.id)
    )
    convs = result.scalars().all()
    unread = any(
        any(
            m.sender == "admin"
            and (c.user_last_read_at is None or m.created_at > c.user_last_read_at)
            for m in c.messages
        )
        for c in convs
    )
    return {"unread": unread}


@fastapi_app.get("/api/admin/support/conversations")
async def admin_list_conversations(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(
        select(SupportConversation)
        .options(selectinload(SupportConversation.messages))
        .order_by(SupportConversation.updated_at.desc())
    )
    return [_conv_to_dict(c) for c in result.scalars().all()]


@fastapi_app.get("/api/admin/support/conversations/{conv_id}")
async def admin_get_conversation(
    conv_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    conv = await _load_conversation_with_messages(db, conv_id)
    if not conv:
        raise HTTPException(404, "Диалог не найден")
    if conv.admin_last_read_at is None or any(
        m.sender == "user" and (conv.admin_last_read_at is None or m.created_at > conv.admin_last_read_at)
        for m in conv.messages
    ):
        conv.admin_last_read_at = datetime.utcnow()
        await db.commit()
    return _conv_to_dict(conv, with_messages=True)


@fastapi_app.post("/api/admin/support/conversations/{conv_id}/messages")
async def admin_reply_message(
    conv_id: int,
    data: SupportMessageCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    text = data.text.strip()
    if not text:
        raise HTTPException(400, "Сообщение не может быть пустым")

    conv = await _load_conversation_with_messages(db, conv_id)
    if not conv:
        raise HTTPException(404, "Диалог не найден")

    msg = SupportMessage(conversation_id=conv.id, sender="admin", text=text, images=data.images or [])
    db.add(msg)
    await db.commit()
    await db.refresh(msg)
    if conv.user_id:
        await sio.emit(
            "support_reply",
            {"conversation_id": conv.id, "message": _msg_to_dict(msg)},
            room=f"user_{conv.user_id}",
        )
    return _msg_to_dict(msg)


@fastapi_app.delete("/api/admin/support/conversations/{conv_id}")
async def admin_delete_conversation(
    conv_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    conv = await _load_conversation_with_messages(db, conv_id)
    if not conv:
        raise HTTPException(404, "Диалог не найден")

    user_id = conv.user_id
    await db.delete(conv)
    await db.commit()
    if user_id:
        await sio.emit(
            "support_conversation_deleted",
            {"conversation_id": conv_id},
            room=f"user_{user_id}",
        )
    return {"ok": True}


@fastapi_app.get("/api/categories")
async def list_categories(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(
            Category.id,
            Category.name,
            select(func.count(Product.id))
            .where(Product.category_id == Category.id)
            .correlate(Category)
            .scalar_subquery().label("product_count"),
        ).order_by(Category.id)
    )
    return [
        {"id": row.id, "name": row.name, "product_count": row.product_count}
        for row in result.all()
    ]


@fastapi_app.post("/api/categories")
async def create_category(data: CategoryCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    cat = Category(name=data.name)
    db.add(cat)
    await db.commit()
    await db.refresh(cat)
    return cat


@fastapi_app.post("/api/products/similar")
async def find_similar_products(data: SimilarRequest, db: AsyncSession = Depends(get_db)):
    color_expr = Product.attributes["color"].as_string()
    result = await db.execute(
        select(Product.id).where(
            Product.id != data.exclude_id,
            Product.name.ilike(data.name),
            color_expr.ilike(data.color),
        )
    )
    return [row[0] for row in result.all()]


@fastapi_app.get("/api/products")
async def list_products(
    category_ids: Optional[str] = None,
    search: Optional[str] = None,
    offset: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
):
    base = select(Product.id)
    total_q = select(func.count(Product.id))

    if category_ids:
        ids_list = [int(x) for x in category_ids.split(",") if x.strip()]
        if ids_list:
            base = base.where(Product.category_id.in_(ids_list))
            total_q = total_q.where(Product.category_id.in_(ids_list))
    if search:
        base = base.where(Product.name.ilike(f"%{search}%"))
        total_q = total_q.where(Product.name.ilike(f"%{search}%"))
    total_result = await db.execute(total_q)
    total = total_result.scalar() or 0

    id_result = await db.execute(
        base.order_by(Product.id.desc()).offset(offset).limit(limit)
    )
    ids = [row[0] for row in id_result.all()]

    items: list[Product] = []
    if ids:
        result = await db.execute(
            select(Product)
            .options(selectinload(Product.images), selectinload(Product.category))
            .where(Product.id.in_(ids))
            .order_by(Product.id.desc())
        )
        items = list(result.scalars().all())

    return {"items": items, "total": total}


@fastapi_app.get("/api/products/{product_id:int}")
async def get_product(product_id: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.images), selectinload(Product.category))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    return product


@fastapi_app.post("/api/products")
async def create_product(data: ProductCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    cat_result = await db.execute(select(Category).where(Category.id == data.category_id))
    if not cat_result.scalar_one_or_none():
        raise HTTPException(400, "Category not found")

    product = Product(
        name=data.name,
        price=data.price,
        purchase_price=data.purchase_price,
        is_available=data.is_available,
        quantity=data.quantity,
        category_id=data.category_id,
        attributes=data.attributes,
    )
    db.add(product)
    await db.flush()

    for img in data.images:
        pi = ProductImage(
            product_id=product.id,
            image_url=img["image_url"],
            is_main=img.get("is_main", False),
        )
        db.add(pi)

    await _upsert_photo_group(db, product.name, _product_color(data.attributes))

    await db.commit()
    await db.refresh(product)
    return product


@fastapi_app.put("/api/products/{product_id:int}")
async def update_product(product_id: int, data: ProductUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    update_data = data.model_dump(exclude_unset=True)
    images = update_data.pop("images", None)

    for key, value in update_data.items():
        setattr(product, key, value)

    if images is not None:
        await db.execute(delete(ProductImage).where(ProductImage.product_id == product_id))
        for img in images:
            pi = ProductImage(
                product_id=product_id,
                image_url=img["image_url"],
                is_main=img.get("is_main", False),
            )
            db.add(pi)

    await _upsert_photo_group(db, product.name, _product_color(product.attributes))

    await db.commit()
    await db.refresh(product)
    return product


@fastapi_app.delete("/api/products/{product_id:int}")
async def delete_product(product_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    result = await db.execute(
        select(Product)
        .options(selectinload(Product.images))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    # Сохраняем URL фото до удаления
    image_urls = [img.image_url for img in product.images]

    await db.delete(product)
    await db.commit()

    # Проверяем после коммита — товар уже удалён, другие запросы видят актуальное состояние
    for url in image_urls:
        await _delete_photo_file_if_unused(db, url)

    return {"ok": True}


def _product_color(attributes: dict | None) -> str:
    """Цвет товара из attributes (для сопоставления с фото-группами)."""
    c = (attributes or {}).get("color")
    return str(c).strip() if c is not None else ""


def _product_effective_price(p: Product) -> float:
    """Цена для покупателя: цена с маржой, либо (если не задана) цена закупки."""
    if p.price is not None and p.price > 0:
        return float(p.price)
    return float(p.purchase_price or 0)


async def _delete_photo_file_if_unused(db: AsyncSession, url: str):
    """Удаляет файл фото с диска, если на его URL больше никто не ссылается
    (ни товары в product_images, ни фото-группы product_photo_groups)."""
    if not url or "/uploads/" not in url:
        return
    count_result = await db.execute(
        select(func.count(ProductImage.id)).where(ProductImage.image_url == url)
    )
    if (count_result.scalar() or 0) > 0:
        return
    for g in (await db.execute(select(ProductPhotoGroup))).scalars().all():
        if url in (g.images or []):
            return
    fname = url.split("/uploads/", 1)[-1].lstrip("/")
    fpath = os.path.join(UPLOAD_DIR, fname)
    if os.path.exists(fpath):
        os.remove(fpath)


async def _verify_upload_files(db: AsyncSession):
    """Проверяет целостность картинок: каждый /uploads/ URL из БД должен
    существовать на диске. Логирует все битые ссылки — защита от «пропавших» фото."""
    urls: set[str] = set()
    for (u,) in (await db.execute(select(ProductImage.image_url))).all():
        if u:
            urls.add(u)
    for images in (await db.execute(select(ProductPhotoGroup.images))).scalars().all():
        for u in (images or []):
            if u:
                urls.add(u)
    for (u,) in (await db.execute(select(Promo.gift_image))).all():
        if u:
            urls.add(u)

    missing = []
    for url in sorted(urls):
        if "/uploads/" not in url:
            continue
        fname = url.split("/uploads/", 1)[-1].lstrip("/")
        if not os.path.exists(os.path.join(UPLOAD_DIR, fname)):
            missing.append(url)

    if missing:
        logger.warning("⚠️ Целостность картинок: %d ссылок в БД без файлов на диске:", len(missing))
        for u in missing:
            logger.warning("   %s", u)
    return missing


async def _upsert_photo_group(db: AsyncSession, name: str, color: str) -> ProductPhotoGroup:
    """Создаёт фото-группу по (наименование, цвет), если её ещё нет."""
    color = (color or "").strip()
    result = await db.execute(
        select(ProductPhotoGroup).where(
            ProductPhotoGroup.name == name,
            ProductPhotoGroup.color == color,
        )
    )
    group = result.scalar_one_or_none()
    if not group:
        group = ProductPhotoGroup(name=name, color=color, images=[])
        db.add(group)
        await db.flush()
    return group


async def _apply_group_photos_to_product(db: AsyncSession, product: Product, group: ProductPhotoGroup):
    """Привязывает фото группы к конкретному товару (пропуская уже привязанные)."""
    urls = [u for u in (group.images or []) if u]
    if not urls:
        return
    result = await db.execute(
        select(ProductImage.image_url).where(ProductImage.product_id == product.id)
    )
    have = {row[0] for row in result.all()}
    for idx, url in enumerate(urls):
        if url in have:
            continue
        db.add(ProductImage(
            product_id=product.id,
            image_url=url,
            is_main=(not have) and idx == 0,
        ))
        have.add(url)


async def _apply_photo_group_to_products(db: AsyncSession, group: ProductPhotoGroup):
    """Привязывает фото группы ко всем товарам с таким же наименованием и цветом."""
    result = await db.execute(select(Product).where(Product.name == group.name))
    for p in result.scalars().all():
        if _product_color(p.attributes) != group.color:
            continue
        await _apply_group_photos_to_product(db, p, group)


async def _remove_photo_group_from_products(db: AsyncSession, group: ProductPhotoGroup, url: str):
    """Отвязывает конкретное фото группы от всех совпадающих товаров."""
    result = await db.execute(select(Product).where(Product.name == group.name))
    for p in result.scalars().all():
        if _product_color(p.attributes) != group.color:
            continue
        await db.execute(delete(ProductImage).where(
            ProductImage.product_id == p.id,
            ProductImage.image_url == url,
        ))


def _serialize_photo_group(group: ProductPhotoGroup, product_count: int = 0) -> dict:
    return {
        "id": group.id,
        "name": group.name,
        "color": group.color,
        "images": group.images or [],
        "product_count": product_count,
        "created_at": group.created_at,
        "updated_at": group.updated_at,
    }


async def _ensure_photo_groups_from_products(db: AsyncSession):
    """Создаёт недостающие фото-группы по (наименование, цвет) из уже имеющихся товаров."""
    existing = {
        (g.name, g.color)
        for g in (await db.execute(select(ProductPhotoGroup.name, ProductPhotoGroup.color))).all()
    }
    seen: set[tuple[str, str]] = set()
    for name, attrs in (await db.execute(select(Product.name, Product.attributes))).all():
        if name is None:
            continue
        key = (name, _product_color(attrs))
        if key in seen or key in existing:
            continue
        seen.add(key)
        db.add(ProductPhotoGroup(name=key[0], color=key[1], images=[]))
    await db.commit()


@fastapi_app.get("/api/products/photo-groups")
async def list_photo_groups(db: AsyncSession = Depends(get_db)):
    await _ensure_photo_groups_from_products(db)
    groups = (await db.execute(
        select(ProductPhotoGroup).order_by(ProductPhotoGroup.name, ProductPhotoGroup.color)
    )).scalars().all()

    cats = {c.id: c.name for c in (await db.execute(select(Category))).scalars().all()}

    by_key: dict[tuple[str, str], dict] = {}
    for name, attrs, category_id in (
        await db.execute(select(Product.name, Product.attributes, Product.category_id))
    ).all():
        key = (name, _product_color(attrs))
        entry = by_key.setdefault(key, {"count": 0, "cat_ids": set()})
        entry["count"] += 1
        if category_id:
            entry["cat_ids"].add(category_id)

    out = []
    for g in groups:
        entry = by_key.get((g.name, g.color), {"count": 0, "cat_ids": set()})
        cat_ids = sorted(entry["cat_ids"])
        out.append({
            **_serialize_photo_group(g, entry["count"]),
            "category_ids": cat_ids,
            "categories": [cats.get(cid, "") for cid in cat_ids],
        })
    return out


class PhotoGroupImagesAdd(BaseModel):
    images: list[str] = []


@fastapi_app.post("/api/products/photo-groups/{group_id}/images")
async def add_photo_group_images(group_id: int, data: PhotoGroupImagesAdd, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    result = await db.execute(select(ProductPhotoGroup).where(ProductPhotoGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(404, "Photo group not found")

    current = group.images or []
    for url in data.images:
        if url and url not in current:
            current.append(url)
    group.images = current

    await _apply_photo_group_to_products(db, group)
    await db.commit()
    await db.refresh(group)
    return _serialize_photo_group(group)


@fastapi_app.delete("/api/products/photo-groups/{group_id}/images")
async def remove_photo_group_image(group_id: int, image_url: str = Query(...), db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    result = await db.execute(select(ProductPhotoGroup).where(ProductPhotoGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(404, "Photo group not found")

    current = group.images or []
    if image_url in current:
        current.remove(image_url)
    group.images = current

    await _remove_photo_group_from_products(db, group, image_url)
    await db.commit()

    # Файл удаляем только если на него больше нигде не ссылаются
    await _delete_photo_file_if_unused(db, image_url)

    return _serialize_photo_group(group)


async def _remove_photo_group_all_images(db: AsyncSession, group: ProductPhotoGroup):
    """Отвязывает все фото группы от всех совпадающих товаров."""
    urls = [u for u in (group.images or []) if u]
    if not urls:
        return
    result = await db.execute(select(Product).where(Product.name == group.name))
    for p in result.scalars().all():
        if _product_color(p.attributes) != group.color:
            continue
        await db.execute(delete(ProductImage).where(
            ProductImage.product_id == p.id,
            ProductImage.image_url.in_(urls),
        ))


@fastapi_app.delete("/api/products/photo-groups/{group_id}")
async def delete_photo_group(group_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    """Удаляет фото-группу целиком: отвязывает все её фото от товаров,
    помечает (name, color) как удалённый и удаляет файлы, на которые больше никто не ссылается."""
    result = await db.execute(select(ProductPhotoGroup).where(ProductPhotoGroup.id == group_id))
    group = result.scalar_one_or_none()
    if not group:
        raise HTTPException(404, "Photo group not found")

    urls = list(group.images or [])
    await _remove_photo_group_all_images(db, group)
    await db.delete(group)
    await db.commit()

    for url in urls:
        await _delete_photo_file_if_unused(db, url)
    return {"ok": True}


@fastapi_app.delete("/api/products/photo-groups")
async def delete_all_photo_groups(db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    """Удаляет все фото-группы и отвязывает все их фото от товаров."""
    groups = (await db.execute(select(ProductPhotoGroup))).scalars().all()
    all_urls: set[str] = set()
    for g in groups:
        all_urls.update(g.images or [])
        await _remove_photo_group_all_images(db, g)
        await db.delete(g)
    await db.commit()

    for url in all_urls:
        await _delete_photo_file_if_unused(db, url)
    return {"ok": True}


@fastapi_app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename or "file")[1].lower() or ".jpg"
    if ext not in ALLOWED_UPLOAD_EXTENSIONS:
        raise HTTPException(400, "Допустимы только изображения (jpg, png, webp, gif)")

    content = await file.read()
    if len(content) > MAX_UPLOAD_SIZE:
        raise HTTPException(413, "Файл слишком большой (максимум 10 МБ)")
    if not _is_allowed_image(content):
        raise HTTPException(400, "Файл не является изображением")

    filename = f"{uuid.uuid4().hex}{ext}"
    path = os.path.join(UPLOAD_DIR, filename)
    with open(path, "wb") as f:
        f.write(content)
    return {"url": f"/uploads/{filename}"}


@fastapi_app.post("/api/products/{product_id:int}/images")
async def add_product_image(product_id: int, data: dict, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")

    pi = ProductImage(
        product_id=product_id,
        image_url=data["image_url"],
        is_main=data.get("is_main", False),
    )
    db.add(pi)
    await db.commit()
    return pi


SHEET_PARSERS = {
    "iPhone": parse_iphone_excel_to_dicts,
    "MacBook": parse_macbook_excel_to_dicts,
    "iPad": parse_ipad_excel_to_dicts,
    "iMac": parse_imac_excel_to_dicts,
    "Apple Watch": parse_watch_excel_to_dicts,
    "AirPods": parse_airpods_excel_to_dicts,
    "Аксессуары Apple": parse_accessories_excel_to_dicts,
    "Samsung": parse_samsung_excel_to_dicts,
    "Игровые приставки": parse_consoles_excel_to_dicts,
    "Dyson": parse_dyson_excel_to_dicts,
    "Xiaomi": parse_xiaomi_excel_to_dicts,
    "POCO": parse_poco_excel_to_dicts,
    "Яндекс Станции": parse_stations_excel_to_dicts,
}


def validate_excel_file(file: UploadFile):
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "File must be an Excel file (.xlsx or .xls)")


async def save_upload_to_tmp(file: UploadFile, prefix: str) -> str:
    ext = os.path.splitext(file.filename or "file")[1] or ".xlsx"
    tmp_path = os.path.join(UPLOAD_DIR, f"{prefix}_{uuid.uuid4().hex}{ext}")
    content = await file.read()
    with open(tmp_path, "wb") as f:
        f.write(content)
    return tmp_path


@fastapi_app.post("/api/products/excel-preview")
async def excel_preview(file: UploadFile = File(...), current_user: User = Depends(require_admin)):
    validate_excel_file(file)
    tmp_path = await save_upload_to_tmp(file, "_preview")

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True)
        sheets = [{"name": name, "can_parse": name in SHEET_PARSERS} for name in wb.sheetnames]
        wb.close()
        return {"sheets": sheets}
    finally:
        os.remove(tmp_path)


@fastapi_app.post("/api/products/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    sheets: str = Form(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    validate_excel_file(file)
    tmp_path = await save_upload_to_tmp(file, "_import")

    selected = [s.strip() for s in sheets.split(",") if s.strip()] if sheets else []
    if not selected:
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            selected = [n for n in wb.sheetnames if n in SHEET_PARSERS]
            wb.close()
        except Exception:
            os.remove(tmp_path)
            raise HTTPException(400, "Не удалось прочитать файл")

    total_imported = 0
    results = []
    imported_cat_ids: list[int] = []

    for sheet_name in selected:
        parser = SHEET_PARSERS.get(sheet_name)
        if not parser:
            continue

        try:
            items = parser(tmp_path)
        except (ValueError, FileNotFoundError):
            continue

        if not items:
            continue

        cat_name = sheet_name

        cat_result = await db.execute(select(Category).where(Category.name.ilike(cat_name)))
        cat = cat_result.scalar_one_or_none()
        if not cat:
            cat = Category(name=cat_name)
            db.add(cat)
            await db.flush()

        updated = 0
        created = 0
        unchanged = 0
        removed = 0
        hidden = 0
        sheet_total = 0

        groups: dict[str, list[dict]] = defaultdict(list)
        for item in items:
            rd = item.get("raw_data", "") or ""
            groups[rd].append(item)

        seen_raw: set[str] = set()
        photo_to_apply: list[tuple[Product, str]] = []

        for raw_data, group_items in groups.items():
            first = group_items[0]
            quantity = len(group_items)
            seen_raw.add(raw_data)

            product = None

            if raw_data:
                existing = await db.execute(
                    select(Product).where(
                        Product.raw_data == raw_data,
                        Product.category_id == cat.id,
                    )
                )
                product = existing.scalar_one_or_none()

            attrs = {
                k: v for k, v in first.items() if k not in ("model", "price", "raw_data")
            }

            if product:
                product.is_available = True
                # Если цена продажи была просто скопирована из закупки прошлым
                # импортом (т.е. админ её ещё не настраивал), сбрасываем её —
                # цену продажи задаёт админ отдельно (наценкой или вручную).
                if product.purchase_price is not None and product.price == product.purchase_price:
                    product.price = None
                changed = False
                if product.name != first["model"]:
                    product.name = first["model"]
                    changed = True
                if product.purchase_price != first["price"]:
                    product.purchase_price = first["price"]
                    changed = True
                if (product.attributes or {}) != attrs:
                    product.attributes = attrs
                    changed = True
                if product.quantity != quantity:
                    product.quantity = quantity
                    changed = True
                if changed:
                    updated += 1
                else:
                    unchanged += 1
            else:
                product = Product(
                    name=first["model"],
                    raw_data=raw_data or None,
                    price=None,
                    purchase_price=first["price"],
                    is_available=True,
                    quantity=quantity,
                    category_id=cat.id,
                    attributes=attrs,
                )
                db.add(product)
                created += 1

            sheet_total += quantity
            total_imported += quantity
            photo_to_apply.append((product, _product_color(attrs)))

        # Создаём/обновляем фото-группы по (наименование, цвет) и привязываем
        # уже загруженные фото группы к товарам. Группы живут отдельно от
        # products, поэтому фото переживают полную перезаливку таблицы.
        await db.flush()
        for p, color in photo_to_apply:
            group = await _upsert_photo_group(db, p.name, color)
            await _apply_group_photos_to_product(db, p, group)

        # Удаляем товары категории, которых нет в загруженном файле:
        # поставщика товара может уже не быть в наличии.
        all_cat_result = await db.execute(
            select(Product).where(Product.category_id == cat.id)
        )
        for p in all_cat_result.scalars().all():
            key = p.raw_data or ""
            if key in seen_raw:
                continue
            try:
                async with db.begin_nested():
                    await db.delete(p)
                removed += 1
            except IntegrityError:
                # Товар фигурирует в заказах — удалять нельзя (RESTRICT).
                # Скрываем из каталога, но сохраняем для истории заказов.
                p.is_available = False
                hidden += 1

        results.append({
            "category": cat_name,
            "created": created,
            "updated": updated,
            "unchanged": unchanged,
            "removed": removed,
            "hidden": hidden,
            "total": sheet_total,
            "category_id": cat.id,
        })
        imported_cat_ids.append(cat.id)

    # Применяем активные наценки к свежим ценам закупки, чтобы правила
    # ценообразования переживали перезаливку таблицы.
    await _apply_active_margins(db, imported_cat_ids, reset_without_rules=False)

    await db.commit()
    os.remove(tmp_path)

    return {"imported": total_imported, "details": results}


### Наценки (правила ценообразования по категориям)

class MarginCreate(BaseModel):
    # "percent" | "fixed"
    margin_type: str = "percent"
    value: float
    target_category_id: int
    active: bool = True


class MarginUpdate(BaseModel):
    margin_type: str | None = None
    value: float | None = None
    target_category_id: int | None = None
    active: bool | None = None


def _margin_to_dict(m: Margin, category_name: str | None = None) -> dict:
    return {
        "id": m.id,
        "margin_type": m.margin_type,
        "value": m.value,
        "target_category_id": m.target_category_id,
        "target_category_name": (
            m.target_category.name if m.target_category else category_name
        ),
        "active": m.active,
        "created_at": m.created_at,
        "updated_at": m.updated_at,
    }


async def _apply_active_margins(
    db: AsyncSession, category_ids: list[int], reset_without_rules: bool = True
) -> dict:
    """Пересчитывает цены продажи товаров указанных категорий по активным
    наценкам. Правила категории применяются последовательно (по id) к цене
    закупки. Если активных правил в категории не осталось, цена продажи
    сбрасывается — покупателю показывается цена закупки.
    reset_without_rules=False оставляет цены в категориях без правил как есть
    (используется при импорте Excel)."""
    result: dict[int, dict] = {}
    if not category_ids:
        return {"updated": 0, "skipped_no_cost": 0}

    rules_rows = (
        await db.execute(
            select(Margin)
            .where(Margin.active == True, Margin.target_category_id.in_(category_ids))  # noqa: E712
            .order_by(Margin.id)
        )
    ).scalars().all()
    rules_by_cat: dict[int, list[Margin]] = defaultdict(list)
    for r in rules_rows:
        rules_by_cat[r.target_category_id].append(r)

    products = (
        await db.execute(select(Product).where(Product.category_id.in_(category_ids)))
    ).scalars().all()

    updated = 0
    skipped_no_cost = 0
    for p in products:
        rules = rules_by_cat.get(p.category_id) or []
        if not rules:
            if not reset_without_rules:
                continue
            # Активных наценок в категории нет — цена продажи не нужна,
            # покупателю показывается цена закупки.
            if p.price is not None:
                p.price = None
                updated += 1
            continue

        base = p.purchase_price
        if base is None or base <= 0:
            skipped_no_cost += 1
            continue

        val = float(base)
        for r in rules:
            if r.margin_type == "fixed":
                val += float(r.value)
            else:
                val *= 1 + float(r.value) / 100
        new_price = round(val)
        if p.price != new_price:
            updated += 1
        p.price = new_price

    result["updated"] = updated
    result["skipped_no_cost"] = skipped_no_cost
    return result


async def _get_margin_loaded(db: AsyncSession, margin_id: int) -> Margin | None:
    result = await db.execute(
        select(Margin)
        .options(selectinload(Margin.target_category))
        .where(Margin.id == margin_id)
    )
    return result.scalar_one_or_none()


def _validate_margin_payload(margin_type: str | None, value: float | None):
    if margin_type is not None and margin_type not in ("percent", "fixed"):
        raise HTTPException(400, "Тип наценки должен быть percent или fixed")
    if value is not None and value <= 0:
        raise HTTPException(400, "Значение наценки должно быть больше нуля")


@fastapi_app.get("/api/admin/margins")
async def list_margins(db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    rows = (
        await db.execute(
            select(Margin).options(selectinload(Margin.target_category)).order_by(Margin.id)
        )
    ).scalars().all()
    cat_ids = [m.target_category_id for m in rows if m.target_category_id is not None]
    counts: dict[int, int] = {}
    if cat_ids:
        cnt_rows = await db.execute(
            select(Product.category_id, func.count(Product.id))
            .where(Product.category_id.in_(cat_ids))
            .group_by(Product.category_id)
        )
        counts = {cid: cnt for cid, cnt in cnt_rows.all()}
    items = []
    for m in rows:
        d = _margin_to_dict(m)
        d["products_count"] = counts.get(m.target_category_id, 0)
        items.append(d)
    return items


@fastapi_app.post("/api/admin/margins")
async def create_margin(data: MarginCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    _validate_margin_payload(data.margin_type, data.value)
    cat = (await db.execute(select(Category).where(Category.id == data.target_category_id))).scalar_one_or_none()
    if not cat:
        raise HTTPException(404, "Категория не найдена")
    dup = (
        await db.execute(select(Margin).where(Margin.target_category_id == data.target_category_id))
    ).scalar_one_or_none()
    if dup:
        raise HTTPException(400, "Для этой категории уже задана наценка — отредактируйте её")
    margin = Margin(
        margin_type=data.margin_type,
        value=data.value,
        target_category_id=data.target_category_id,
        active=data.active,
    )
    db.add(margin)
    stats = await _apply_active_margins(db, [data.target_category_id])
    await db.commit()
    margin = await _get_margin_loaded(db, margin.id)
    d = _margin_to_dict(margin)
    d["applied"] = stats
    return d


@fastapi_app.patch("/api/admin/margins/{margin_id}")
async def update_margin(margin_id: int, data: MarginUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    margin = (await db.execute(select(Margin).where(Margin.id == margin_id))).scalar_one_or_none()
    if not margin:
        raise HTTPException(status_code=404, detail="Наценка не найдена")
    payload = data.model_dump(exclude_unset=True)
    _validate_margin_payload(payload.get("margin_type"), payload.get("value"))

    affected_cats = {margin.target_category_id}
    if "target_category_id" in payload:
        new_cat_id = payload["target_category_id"]
        cat = (await db.execute(select(Category).where(Category.id == new_cat_id))).scalar_one_or_none()
        if not cat:
            raise HTTPException(404, "Категория не найдена")
        if new_cat_id != margin.target_category_id:
            dup = (
                await db.execute(
                    select(Margin).where(
                        Margin.target_category_id == new_cat_id,
                        Margin.id != margin_id,
                    )
                )
            ).scalar_one_or_none()
            if dup:
                raise HTTPException(400, "Для этой категории уже задана наценка — отредактируйте её")
        affected_cats.add(new_cat_id)
    for key, value in payload.items():
        setattr(margin, key, value)

    stats = await _apply_active_margins(db, list(affected_cats))
    await db.commit()
    margin = await _get_margin_loaded(db, margin_id)
    d = _margin_to_dict(margin)
    d["applied"] = stats
    return d


@fastapi_app.delete("/api/admin/margins/{margin_id}")
async def delete_margin(margin_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    margin = (await db.execute(select(Margin).where(Margin.id == margin_id))).scalar_one_or_none()
    if not margin:
        raise HTTPException(status_code=404, detail="Наценка не найдена")
    cat_id = margin.target_category_id
    await db.delete(margin)
    stats = await _apply_active_margins(db, [cat_id])
    await db.commit()
    return {"ok": True, "applied": stats}


class OrderItemCreate(BaseModel):
    product_id: int
    quantity: int = 1
    price_at_purchase: float
    selected_attributes: dict | None = None


def _order_to_dict(o: Order) -> dict:
    return {
        "id": o.id,
        "user_id": o.user_id,
        "courier_id": o.courier_id,
        "status": o.status.value if hasattr(o.status, "value") else o.status,
        "total_price": o.total_price,
        "delivery_info": o.delivery_info,
        "customer_name": o.customer_name,
        "phone": o.phone,
        "delivery_lat": o.delivery_lat,
        "delivery_lng": o.delivery_lng,
        "created_at": o.created_at,
        "updated_at": o.updated_at,
        "confirmation_code": o.confirmation_code,
        "delivery_imei": o.delivery_imei,
        "delivery_photo_urls": o.delivery_photo_urls or [],
        "delivered_at": o.delivered_at,
        "archived_at": o.archived_at,
        "items": o.items,
        "user_email": o.user.email if o.user else None,
        "user_username": o.user.username if o.user else None,
        "courier_name": o.courier.name if o.courier else None,
        "courier_phone": o.courier.phone if o.courier else None,
        "courier_login": o.courier.login if o.courier else None,
        "gifts": o.gifts or [],
        "trade_in": False,
        "trade_in_description": None,
        "trade_in_photos": [],
        "trade_in_price": None,
    }


def _can_view_order(order: Order, user: User | None) -> bool:
    """Доступ к заказу: админ — ко всем, владелец — к своим, гость — только к гостевым."""
    if user and user.role == UserRole.ADMIN:
        return True
    if order.user_id is None:
        return True
    return bool(user) and order.user_id == user.id


class OrderCreate(BaseModel):
    items: list[OrderItemCreate]
    total_price: float
    customer_name: str | None = None
    delivery_info: str | None = None
    delivery_lat: float | None = None
    delivery_lng: float | None = None
    phone: str | None = None
    # Подарки, которые покупатель согласился получить; None — сохранять все
    accepted_gift_names: list[str] | None = None


class PromoCreate(BaseModel):
    gift_name: str
    gift_image: str | None = None
    gift_price: float = 0
    # "product" | "category" | "all"
    target_type: str = "all"
    target_product_id: int | None = None
    target_category_id: int | None = None
    min_total: float | None = None
    active: bool = True


class PromoUpdate(BaseModel):
    gift_name: str | None = None
    gift_image: str | None = None
    gift_price: float | None = None
    target_type: str | None = None
    target_product_id: int | None = None
    target_category_id: int | None = None
    min_total: float | None = None
    active: bool | None = None


class OrderStatusUpdate(BaseModel):
    status: str


@fastapi_app.post("/api/orders")
async def create_order(data: OrderCreate, db: AsyncSession = Depends(get_db), user: User = Depends(get_optional_user)):
    promos = await _find_promos_for_cart(db, data)
    all_gifts = [
        {"name": p.gift_name, "image": p.gift_image, "price": p.gift_price}
        for p in promos
    ]
    if data.accepted_gift_names is not None:
        accepted = set(data.accepted_gift_names)
        gifts = [g for g in all_gifts if g["name"] in accepted]
    else:
        gifts = all_gifts
    order = Order(
        total_price=data.total_price,
        customer_name=data.customer_name,
        delivery_info=data.delivery_info,
        delivery_lat=data.delivery_lat,
        delivery_lng=data.delivery_lng,
        phone=data.phone,
        user_id=user.id if user else None,
        gifts=gifts if gifts else None,
        status_changed_at=datetime.utcnow(),
    )
    db.add(order)
    await db.flush()
    if user:
        await db.execute(
            update(User).where(User.id == user.id).values(last_orders_seen_at=datetime.utcnow())
        )

    product_ids = [it.product_id for it in data.items]
    products_map: dict[int, Product] = {}
    if product_ids:
        pres = await db.execute(select(Product).options(selectinload(Product.images)).where(Product.id.in_(product_ids)))
        products_map = {p.id: p for p in pres.scalars().all()}

    for item in data.items:
        product = products_map.get(item.product_id)
        main_img = None
        if product and product.images:
            main_img = next((img.image_url for img in product.images if img.is_main), product.images[0].image_url)
        order_item = OrderItem(
            order_id=order.id,
            product_id=item.product_id,
            quantity=item.quantity,
            price_at_purchase=item.price_at_purchase,
            selected_attributes=item.selected_attributes,
            product_name=product.name if product else None,
            product_image=main_img,
        )
        db.add(order_item)

    await db.commit()

    result = await db.execute(
        select(Order)
        .options(
            selectinload(Order.items).selectinload(OrderItem.product).selectinload(Product.images),
            selectinload(Order.user),
        )
        .where(Order.id == order.id)
    )
    o = result.scalar_one()
    return _order_to_dict(o)


### Акции / подарки

def _promo_to_dict(p: Promo) -> dict:
    return {
        "id": p.id,
        "gift_name": p.gift_name,
        "gift_image": p.gift_image,
        "gift_price": p.gift_price,
        "target_type": p.target_type,
        "target_product_id": p.target_product_id,
        "target_product_name": p.target_product.name if p.target_product else None,
        "target_category_id": p.target_category_id,
        "target_category_name": p.target_category.name if p.target_category else None,
        "min_total": p.min_total,
        "active": p.active,
        "created_at": p.created_at,
        "updated_at": p.updated_at,
    }


async def _get_promo_loaded(db: AsyncSession, promo_id: int) -> Promo | None:
    """Загружает промо с отношениями (target_product/target_category), чтобы
    избежать lazy-load в async-контексте при сериализации."""
    result = await db.execute(
        select(Promo)
        .options(selectinload(Promo.target_product), selectinload(Promo.target_category))
        .where(Promo.id == promo_id)
    )
    return result.scalar_one_or_none()


async def _find_promos_for_cart(db: AsyncSession, data: OrderCreate) -> list[Promo]:
    """Возвращает все активные промо, чьи условия выполняются для корзины."""
    promos = (
        await db.execute(
            select(Promo)
            .options(selectinload(Promo.target_product), selectinload(Promo.target_category))
            .where(Promo.active == True)  # noqa: E712
            .order_by(Promo.id)
        )
    ).scalars().all()
    if not promos:
        return []

    product_ids = [it.product_id for it in data.items]
    products: dict[int, Product] = {}
    if product_ids:
        rows = (await db.execute(select(Product).where(Product.id.in_(product_ids)))).scalars().all()
        products = {p.id: p for p in rows}

    total = sum((_product_effective_price(products.get(it.product_id)) if products.get(it.product_id) else it.price_at_purchase) * it.quantity for it in data.items)

    def matches(p: Promo) -> bool:
        if p.min_total is not None and total < p.min_total:
            return False
        if p.target_type == "product":
            return p.target_product_id in product_ids
        if p.target_type == "category":
            if p.target_category_id is None:
                return False
            return any(prod.category_id == p.target_category_id for prod in products.values())
        return True

    return [p for p in promos if matches(p)]


@fastapi_app.get("/api/promos")
async def list_promos(db: AsyncSession = Depends(get_db)):
    rows = (
        await db.execute(
            select(Promo)
            .options(selectinload(Promo.target_product), selectinload(Promo.target_category))
            .order_by(Promo.id)
        )
    ).scalars().all()
    return [_promo_to_dict(p) for p in rows]


@fastapi_app.post("/api/promos")
async def create_promo(data: PromoCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    promo = Promo(
        gift_name=data.gift_name,
        gift_image=data.gift_image,
        gift_price=data.gift_price,
        target_type=data.target_type,
        target_product_id=data.target_product_id,
        target_category_id=data.target_category_id,
        min_total=data.min_total,
        active=data.active,
    )
    db.add(promo)
    await db.commit()
    promo = await _get_promo_loaded(db, promo.id)
    return _promo_to_dict(promo)


@fastapi_app.patch("/api/promos/{promo_id}")
async def update_promo(promo_id: int, data: PromoUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    promo = (await db.execute(select(Promo).where(Promo.id == promo_id))).scalar_one_or_none()
    if not promo:
        raise HTTPException(status_code=404, detail="Промо не найдено")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(promo, key, value)
    await db.commit()
    promo = await _get_promo_loaded(db, promo.id)
    return _promo_to_dict(promo)


@fastapi_app.delete("/api/promos/{promo_id}")
async def delete_promo(promo_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    promo = (await db.execute(select(Promo).where(Promo.id == promo_id))).scalar_one_or_none()
    if not promo:
        raise HTTPException(status_code=404, detail="Промо не найдено")
    await db.delete(promo)
    await db.commit()
    return {"ok": True}


@fastapi_app.post("/api/promos/compute")
async def compute_promo(data: OrderCreate, db: AsyncSession = Depends(get_db)):
    """Вычисляет подарки для корзины (используется фронтендом в оформлении заказа)."""
    promos = await _find_promos_for_cart(db, data)
    return {
        "gifts": [
            {"name": p.gift_name, "image": p.gift_image, "price": p.gift_price}
            for p in promos
        ]
    }

@fastapi_app.get("/api/orders")
async def list_orders(
    status: str | None = None,
    offset: int = 0,
    limit: int = 50,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_optional_user),
):
    base = select(Order)
    count_q = select(func.count(Order.id))

    if status:
        base = base.where(Order.status == status)
        count_q = count_q.where(Order.status == status)

    if not user or user.role != UserRole.ADMIN:
        if user:
            base = base.where(Order.user_id == user.id)
            count_q = count_q.where(Order.user_id == user.id)
        else:
            base = base.where(Order.user_id == None)
            count_q = count_q.where(Order.user_id == None)

    total_result = await db.execute(count_q)
    total = total_result.scalar() or 0

    result = await db.execute(
        base.options(
            selectinload(Order.items).selectinload(OrderItem.product).selectinload(Product.images),
            selectinload(Order.user),
        )
        .order_by(Order.id.desc())
        .offset(offset)
        .limit(limit)
    )
    orders = list(result.scalars().unique().all())

    items = [_order_to_dict(o) for o in orders]

    return {"items": items, "total": total}


@fastapi_app.get("/api/orders/unread")
async def orders_unread(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not current_user.last_orders_seen_at:
        return {"unread": False}
    result = await db.execute(
        select(func.count(Order.id)).where(
            Order.user_id == current_user.id,
            Order.status != OrderStatus.CREATED,
            Order.status_changed_at > current_user.last_orders_seen_at,
        )
    )
    return {"unread": (result.scalar() or 0) > 0}


@fastapi_app.post("/api/orders/read")
async def mark_orders_read(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await db.execute(
        update(User).where(User.id == current_user.id).values(last_orders_seen_at=datetime.utcnow())
    )
    await db.commit()
    return {"ok": True}


@fastapi_app.get("/api/orders/bulk")
async def get_orders_by_ids(ids: str, db: AsyncSession = Depends(get_db), user: User = Depends(get_optional_user)):
    id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
    if not id_list:
        return []
    result = await db.execute(
        select(Order)
        .options(
            selectinload(Order.items).selectinload(OrderItem.product).selectinload(Product.images),
            selectinload(Order.user),
        )
        .where(Order.id.in_(id_list))
        .order_by(Order.id.desc())
    )
    orders = list(result.scalars().unique().all())
    return [_order_to_dict(o) for o in orders if _can_view_order(o, user)]


@fastapi_app.get("/api/orders/{order_id}")
async def get_order(order_id: int, db: AsyncSession = Depends(get_db), user: User = Depends(get_optional_user)):
    result = await db.execute(
        select(Order)
        .options(
            selectinload(Order.items).selectinload(OrderItem.product).selectinload(Product.images),
            selectinload(Order.user),
        )
        .where(Order.id == order_id)
    )
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Order not found")
    if not _can_view_order(order, user):
        raise HTTPException(404, "Order not found")
    return _order_to_dict(order)


async def _notify_order_status(order, status: str):
    if order.user_id:
        await sio.emit("order_status_updated", {
            "order_id": order.id,
            "status": status,
        }, room=f"user_{order.user_id}")


async def _get_courier_for_user(db: AsyncSession, user_id: int) -> Courier | None:
    """Профиль курьера по id пользователя (для проверки назначения на заказ)."""
    result = await db.execute(
        select(User).where(User.id == user_id, User.role == UserRole.COURIER)
    )
    user = result.scalar_one_or_none()
    if not user or not user.username:
        return None
    result = await db.execute(select(Courier).where(Courier.login == user.username))
    return result.scalar_one_or_none()


@fastapi_app.patch("/api/orders/{order_id}")
async def update_order_status(order_id: int, data: OrderStatusUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Order not found")

    try:
        new_status = OrderStatus(data.status)
    except ValueError:
        raise HTTPException(400, "Недопустимый статус")

    if current_user.role == UserRole.ADMIN:
        pass
    elif current_user.role == UserRole.COURIER:
        courier = await _get_courier_for_user(db, current_user.id)
        if not courier or order.courier_id != courier.id:
            raise HTTPException(403, "Заказ не назначен вам")
        if new_status not in (OrderStatus.SHIPPED, OrderStatus.COMPLETED):
            raise HTTPException(403, "Недостаточно прав для этого статуса")
    else:
        if order.user_id != current_user.id or new_status != OrderStatus.CANCELLED:
            raise HTTPException(403, "Forbidden")

    order.status = new_status
    order.status_changed_at = datetime.utcnow()
    if new_status == OrderStatus.COMPLETED and order.delivered_at is None:
        order.delivered_at = datetime.utcnow()
        order.archived_at = datetime.utcnow()
    elif new_status != OrderStatus.COMPLETED:
        order.delivered_at = None
        order.archived_at = None
    await db.commit()

    result = await db.execute(
        select(Order)
        .options(
            selectinload(Order.items).selectinload(OrderItem.product).selectinload(Product.images),
            selectinload(Order.user),
        )
        .where(Order.id == order_id)
    )
    o = result.scalar_one()

    await sio.emit("order_updated", {
        "order_id": order_id,
        "status": new_status.value,
    }, room="admin_room")
    await _notify_order_status(o, new_status.value)

    return _order_to_dict(o)


@fastapi_app.delete("/api/orders/{order_id}")
async def delete_order(order_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(require_admin)):
    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Order not found")
    await db.delete(order)
    await db.commit()
    return {"ok": True}


# ==========================================
# COURIER CRUD
# ==========================================

class CourierCreate(BaseModel):
    name: str
    phone: str | None = None


class CourierUpdate(BaseModel):
    name: str | None = None
    phone: str | None = None


class AssignCourierRequest(BaseModel):
    courier_id: int


class CompleteDeliveryRequest(BaseModel):
    confirmation_code: str
    imei: str
    photo_urls: list[str] = []


class VerifyCodeRequest(BaseModel):
    confirmation_code: str


def _generate_courier_login() -> str:
    return "courier_" + uuid.uuid4().hex[:8]


def _generate_courier_password() -> str:
    import secrets
    chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789!@#"
    return "".join(secrets.choice(chars) for _ in range(10))


@fastapi_app.post("/api/couriers")
async def create_courier(data: CourierCreate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    login = _generate_courier_login()
    # Ensure uniqueness
    while True:
        existing = await db.execute(select(Courier).where(Courier.login == login))
        if not existing.scalar_one_or_none():
            break
        login = _generate_courier_login()

    password = _generate_courier_password()
    courier = Courier(
        name=data.name.strip(),
        phone=data.phone.strip() if data.phone else None,
        login=login,
        password_hash=hash_password(password),
    )
    db.add(courier)
    await db.flush()

    user = User(
        name=data.name.strip(),
        phone=data.phone.strip() if data.phone else None,
        username=login,
        role=UserRole.COURIER,
    )
    db.add(user)
    await db.flush()

    identity = UserIdentity(
        user_id=user.id,
        provider=AuthProvider.LOCAL,
        provider_user_id=login,
        password_hash=hash_password(password),
    )
    db.add(identity)
    await db.commit()
    await db.refresh(courier)

    return {
        "id": courier.id,
        "name": courier.name,
        "phone": courier.phone,
        "login": courier.login,
        "password": password,
        "created_at": courier.created_at.isoformat() if courier.created_at else None,
    }


@fastapi_app.get("/api/couriers")
async def list_couriers(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Courier).order_by(Courier.id.desc()))
    couriers = result.scalars().all()
    return [
        {
            "id": c.id,
            "name": c.name,
            "phone": c.phone,
            "login": c.login,
            "password": None,
            "created_at": c.created_at.isoformat() if c.created_at else None,
        }
        for c in couriers
    ]


@fastapi_app.patch("/api/couriers/{courier_id}")
async def update_courier(courier_id: int, data: CourierUpdate, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Courier).where(Courier.id == courier_id))
    courier = result.scalar_one_or_none()
    if not courier:
        raise HTTPException(404, "Courier not found")

    if data.name is not None:
        courier.name = data.name.strip()
    if data.phone is not None:
        courier.phone = data.phone.strip() if data.phone.strip() else None

    user_result = await db.execute(select(User).where(User.username == courier.login))
    courier_user = user_result.scalar_one_or_none()
    if courier_user:
        if data.name is not None:
            courier_user.name = courier.name
        if data.phone is not None:
            courier_user.phone = courier.phone

    await db.commit()
    await db.refresh(courier)
    return {
        "id": courier.id,
        "name": courier.name,
        "phone": courier.phone,
        "login": courier.login,
        "password": None,
        "created_at": courier.created_at.isoformat() if courier.created_at else None,
    }


@fastapi_app.delete("/api/couriers/{courier_id}")
async def delete_courier(courier_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Courier).where(Courier.id == courier_id))
    courier = result.scalar_one_or_none()
    if not courier:
        raise HTTPException(404, "Courier not found")

    user_result = await db.execute(select(User).where(User.username == courier.login))
    courier_user = user_result.scalar_one_or_none()

    orders_result = await db.execute(select(Order).where(Order.courier_id == courier_id))
    for order in orders_result.scalars().all():
        order.courier_id = None

    await db.delete(courier)
    if courier_user:
        await db.delete(courier_user)
    await db.commit()
    return {"ok": True}


@fastapi_app.patch("/api/orders/{order_id}/courier")
async def assign_courier_to_order(order_id: int, data: AssignCourierRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.ADMIN:
        raise HTTPException(403, "Admin only")

    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Order not found")

    # Verify courier exists
    courier_result = await db.execute(select(Courier).where(Courier.id == data.courier_id))
    if not courier_result.scalar_one_or_none():
        raise HTTPException(404, "Courier not found")

    import random
    order.courier_id = data.courier_id
    order.status = OrderStatus.SHIPPED
    order.status_changed_at = datetime.utcnow()
    order.confirmation_code = str(random.randint(10000, 99999))
    await db.commit()

    await sio.emit("order_updated", {
        "order_id": order_id,
        "status": "shipped",
    }, room="admin_room")
    await _notify_order_status(order, "shipped")

    return {"ok": True, "order_id": order_id, "courier_id": data.courier_id}


@fastapi_app.post("/api/orders/{order_id}/verify-code")
async def verify_confirmation_code(order_id: int, data: VerifyCodeRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.COURIER:
        raise HTTPException(403, "Только курьер может подтверждать код")

    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")

    courier = await _get_courier_for_user(db, current_user.id)
    if not courier or order.courier_id != courier.id:
        raise HTTPException(403, "Заказ не назначен вам")

    if order.confirmation_code != data.confirmation_code:
        raise HTTPException(400, "Неверный код подтверждения")

    return {"ok": True, "order_id": order_id}


@fastapi_app.post("/api/orders/{order_id}/complete-delivery")
async def complete_delivery(order_id: int, data: CompleteDeliveryRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.COURIER:
        raise HTTPException(403, "Только курьер может завершить доставку")

    result = await db.execute(select(Order).where(Order.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(404, "Заказ не найден")

    courier = await _get_courier_for_user(db, current_user.id)
    if not courier or order.courier_id != courier.id:
        raise HTTPException(403, "Заказ не назначен вам")

    if order.status == "completed":
        raise HTTPException(400, "Заказ уже доставлен")

    if order.confirmation_code != data.confirmation_code:
        raise HTTPException(400, "Неверный код подтверждения")

    order.delivery_imei = data.imei
    order.delivery_photo_urls = data.photo_urls
    order.status = OrderStatus.COMPLETED
    order.status_changed_at = datetime.utcnow()
    order.delivered_at = datetime.utcnow()
    order.archived_at = datetime.utcnow()
    await db.commit()

    await sio.emit("order_updated", {
        "order_id": order_id,
        "status": "completed",
        "delivered_at": order.delivered_at.isoformat(),
    }, room="admin_room")
    await _notify_order_status(order, OrderStatus.COMPLETED.value)

    return {"ok": True, "order_id": order_id, "delivered_at": order.delivered_at.isoformat()}


@fastapi_app.get("/api/couriers/me/orders")
async def get_courier_my_orders(db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != UserRole.COURIER:
        raise HTTPException(403, "Только курьер")
    courier_result = await db.execute(select(Courier).where(Courier.login == current_user.username))
    courier = courier_result.scalar_one_or_none()
    if not courier:
        raise HTTPException(404, "Профиль курьера не найден")

    result = await db.execute(
        select(Order)
        .where(Order.courier_id == courier.id)
        .options(
            selectinload(Order.items).selectinload(OrderItem.product).selectinload(Product.images),
            selectinload(Order.user),
        )
        .order_by(Order.id.desc())
    )
    orders = list(result.scalars().unique().all())

    return [_order_to_dict(o) for o in orders]


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
