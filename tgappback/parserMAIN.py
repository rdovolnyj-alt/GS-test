import os
import re
import openpyxl

COUNTRY_FLAGS = {
    "IN": "🇮🇳 IN",
    "India": "🇮🇳 IN",
    "US": "🇺🇸 US",
    "USA": "🇺🇸 US",
    "CN": "🇨🇳 CN",
    "China": "🇨🇳 CN",
    "AE": "🇦🇪 AE",
    "UAE": "🇦🇪 AE",
    "RU": "🇷🇺 RU",
    "Russia": "🇷🇺 RU"
}


def parse_price(price_raw):
    try:
        if isinstance(price_raw, str):
            return float(price_raw.replace(".", ""))
        return float(price_raw)
    except (ValueError, TypeError):
        return None


def resolve_country(country_raw):
    country_clean = str(country_raw).strip() if country_raw else "Не указана"
    return COUNTRY_FLAGS.get(country_clean, country_clean)


# ── iPhone ──────────────────────────────────────────────


def parse_model_string(raw_text):
    match = re.search(r'(.*?)\s+(\d+(?:GB|TB))\s+(.*)', raw_text, re.IGNORECASE)

    if match:
        model_part = match.group(1).strip()
        memory = match.group(2).strip()
        color = match.group(3).strip()

        if not model_part.lower().startswith('iphone'):
            phone_model = f"iPhone {model_part}"
        else:
            phone_model = model_part

        return phone_model, memory, color
    else:
        return f"iPhone {raw_text}", "Не указано", "Не указано"


def parse_iphone_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "iPhone" not in workbook.sheetnames:
        raise ValueError("Лист 'iPhone' не найден в файле.")

    sheet = workbook["iPhone"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                phone_model, memory, color = parse_model_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                country = resolve_country(country_raw)

                results.append({
                    "raw_data": f"{raw_text} | {country}",
                    "model": phone_model,
                    "memory": memory,
                    "color": color,
                    "price": price,
                    "country": country,
                })

    return results


# ── MacBook ─────────────────────────────────────────────


def parse_macbook_string(raw_text):
    colors_multi = r'Space\s+(?:Black|Gray|Grey)|Sky\s+Blue|Midnight|Starlight|Rose\s+Gold|Product\s+RED|Deep\s+Purple|Sierra\s+Blue|Alpine\s+Green'
    colors_single = r'Silver|Gold|Graphite|Blue|Green|Pink|Purple|Red|White|Black|Grey|Gray|Yellow|Orange|Coral'
    color_pat = rf'({colors_multi}|{colors_single})'

    sto = r'(\d+(?:GB|TB))'
    pats = [
        rf'^(.+?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d{{4}})\s+(\d+)/{sto}\s+{color_pat}',
        rf'^(.+?)\s+(\d{{4}})\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d+)/{sto}\s+{color_pat}',
        rf'^(.+?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d{{4}})\s+(\d+)/{sto}',
        rf'^(.+?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d+)/{sto}\s+{color_pat}',
        rf'^(.+?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d+)/{sto}',
        rf'^(.+?)\s+((?:Core\s+)?i[3579](?:\s+\w+)?)\s+(\d{{4}})\s+(\d+)/{sto}\s+{color_pat}',
        rf'^(.+?)\s+(\d{{4}})\s+((?:Core\s+)?i[3579](?:\s+\w+)?)\s+(\d+)/{sto}\s+{color_pat}',
        rf'^(.+?)\s+(\d{{4}})\s+(\d+)/{sto}\s+{color_pat}',
        rf'^(.+?)\s+(\d{{4}})\s+(\d+)/{sto}',
        rf'^(.+?)\s+(\d+)/{sto}\s+{color_pat}',
        rf'^(.+?)\s+(\d+)/{sto}',
    ]

    for i, pat in enumerate(pats):
        m = re.search(pat, raw_text, re.IGNORECASE)
        if not m:
            continue

        if i == 0:
            return (m.group(1).strip(), m.group(2).strip(), m.group(3).strip(),
                    m.group(4).strip(), m.group(5).strip(), m.group(6).strip())
        if i == 1:
            return (m.group(1).strip(), m.group(3).strip(), m.group(2).strip(),
                    m.group(4).strip(), m.group(5).strip(), m.group(6).strip())
        if i == 2:
            return (m.group(1).strip(), m.group(2).strip(), m.group(3).strip(),
                    m.group(4).strip(), m.group(5).strip(), "Не указано")
        if i == 3:
            return (m.group(1).strip(), m.group(2).strip(), "Не указано",
                    m.group(3).strip(), m.group(4).strip(), m.group(5).strip())
        if i == 4:
            return (m.group(1).strip(), m.group(2).strip(), "Не указано",
                    m.group(3).strip(), m.group(4).strip(), "Не указано")
        if i == 5:
            return (m.group(1).strip(), m.group(2).strip(), m.group(3).strip(),
                    m.group(4).strip(), m.group(5).strip(), m.group(6).strip())
        if i == 6:
            return (m.group(1).strip(), m.group(3).strip(), m.group(2).strip(),
                    m.group(4).strip(), m.group(5).strip(), m.group(6).strip())
        if i == 7:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    m.group(3).strip(), m.group(4).strip(), m.group(5).strip())
        if i == 8:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    m.group(3).strip(), m.group(4).strip(), "Не указано")
        if i == 9:
            return (m.group(1).strip(), "Не указано", "Не указано",
                    m.group(2).strip(), m.group(3).strip(), m.group(4).strip())
        if i == 10:
            return (m.group(1).strip(), "Не указано", "Не указано",
                    m.group(2).strip(), m.group(3).strip(), "Не указано")

    return (raw_text, "Не указано", "Не указано", "Не указано", "Не указано", "Не указано")


def parse_macbook_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "MacBook" not in workbook.sheetnames:
        raise ValueError("Лист 'MacBook' не найден в файле.")

    sheet = workbook["MacBook"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                variant, processor, year, memory, storage, color = parse_macbook_string(raw_text)
                if variant.startswith(("MacBook ", "Mac ")):
                    full_model = variant
                else:
                    full_model = f"MacBook {variant}"

                price = parse_price(price_raw)
                if price is None:
                    continue

                country = resolve_country(country_raw)

                results.append({
                    "raw_data": f"{raw_text} | {country}",
                    "model": full_model,
                    "year": year,
                    "processor": processor,
                    "memory": memory,
                    "storage": storage,
                    "color": color,
                    "price": price,
                    "country": country,
                })

    return results


# ── iPad ────────────────────────────────────────────────

COLORS_MULTI = r'Space\s+(?:Black|Gray|Grey)|Sky\s+Blue|Midnight|Starlight|Rose\s+Gold|Product\s+RED|Deep\s+Purple|Sierra\s+Blue|Alpine\s+Green'
COLORS_SINGLE = r'Silver|Gold|Graphite|Blue|Green|Pink|Purple|Red|White|Black|Grey|Gray|Yellow|Orange|Coral'
COLOR_PAT = rf'({COLORS_MULTI}|{COLORS_SINGLE})'

CONNECTIVITY = r'(Wi-Fi|LTE|5G|Cellular)'

IPAD_PATTERNS = [
    # model chip year storage connectivity color
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d{{4}})\s+(\d+(?:GB|TB))\s+{CONNECTIVITY}\s+{COLOR_PAT}',
    # model chip storage connectivity color
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d+(?:GB|TB))\s+{CONNECTIVITY}\s+{COLOR_PAT}',
    # model chip storage color
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d+(?:GB|TB))\s+{COLOR_PAT}',
    # model chip storage only
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+(\d+(?:GB|TB))',
    # model storage connectivity color (no chip)
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(\d+(?:GB|TB))\s+{CONNECTIVITY}\s+{COLOR_PAT}',
    # model storage connectivity only (no chip, no color)
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(\d+(?:GB|TB))\s+{CONNECTIVITY}',
    # model storage color (no chip)
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(\d+(?:GB|TB))\s+{COLOR_PAT}',
    # model storage only (no chip)
    rf'^(iPad\s+\w+\s+\d+(?:\.\d+)?)\s+(\d+(?:GB|TB))',
    # "iPad 10 64GB Wi-Fi Silver" — variant is just a number
    rf'^(iPad\s+\d+(?:\.\d+)?)\s+(\d+(?:GB|TB))\s+{CONNECTIVITY}\s+{COLOR_PAT}',
    rf'^(iPad\s+\d+(?:\.\d+)?)\s+(\d+(?:GB|TB))\s+{COLOR_PAT}',
    rf'^(iPad\s+\d+(?:\.\d+)?)\s+(\d+(?:GB|TB))',
    # last resort — any "iPad ..." + storage
    rf'^(iPad\s+.+?)\s+(\d+(?:GB|TB))\s+{CONNECTIVITY}\s+{COLOR_PAT}',
    rf'^(iPad\s+.+?)\s+(\d+(?:GB|TB))\s+{COLOR_PAT}',
    rf'^(iPad\s+.+?)\s+(\d+(?:GB|TB))',
]


def parse_ipad_string(raw_text):
    """
    Разбирает строку с моделью iPad.
    Возвращает (model, processor, memory, connectivity, color).
    """
    for i, pat in enumerate(IPAD_PATTERNS):
        m = re.search(pat, raw_text, re.IGNORECASE)
        if not m:
            continue

        # 0: model chip year storage connectivity color
        if i == 0:
            return (m.group(1).strip(), m.group(2).strip(), m.group(3).strip(),
                    m.group(4).strip(), m.group(5).strip())
        # 1: model chip storage connectivity color
        if i == 1:
            return (m.group(1).strip(), m.group(2).strip(), m.group(3).strip(),
                    m.group(4).strip(), m.group(5).strip())
        # 2: model chip storage color
        if i == 2:
            return (m.group(1).strip(), m.group(2).strip(), m.group(3).strip(),
                    "Не указано", m.group(4).strip())
        # 3: model chip storage only
        if i == 3:
            return (m.group(1).strip(), m.group(2).strip(), m.group(3).strip(),
                    "Не указано", "Не указано")
        # 4: model storage connectivity color (no chip)
        if i == 4:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    m.group(3).strip(), m.group(4).strip())
        # 5: model storage connectivity only (no chip, no color)
        if i == 5:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    m.group(3).strip(), "Не указано")
        # 6: model storage color (no chip, no connectivity)
        if i == 6:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    "Не указано", m.group(3).strip())
        # 7: model storage only
        if i == 7:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    "Не указано", "Не указано")
        # 8: "iPad 10 64GB Wi-Fi Silver"
        if i == 8:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    m.group(3).strip(), m.group(4).strip())
        # 9: "iPad 10 64GB Silver"
        if i == 9:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    "Не указано", m.group(3).strip())
        # 10: "iPad 10 64GB"
        if i == 10:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    "Не указано", "Не указано")
        # 11: last resort with connectivity + color
        if i == 11:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    m.group(3).strip(), m.group(4).strip())
        # 12: last resort with color
        if i == 12:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    "Не указано", m.group(3).strip())
        # 13: last resort
        if i == 13:
            return (m.group(1).strip(), "Не указано", m.group(2).strip(),
                    "Не указано", "Не указано")

    return (raw_text, "Не указано", "Не указано", "Не указано", "Не указано")


def parse_ipad_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "iPad" not in workbook.sheetnames:
        raise ValueError("Лист 'iPad' не найден в файле.")

    sheet = workbook["iPad"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model, processor, memory, connectivity, color = parse_ipad_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                country = resolve_country(country_raw)

                results.append({
                    "raw_data": f"{raw_text} | {country}",
                    "model": model,
                    "processor": processor,
                    "memory": memory,
                    "connectivity": connectivity,
                    "color": color,
                    "price": price,
                    "country": country,
                })

    return results


# ── Аксессуары Apple ────────────────────────────────────


def parse_accessories_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "Аксессуары Apple" not in workbook.sheetnames:
        raise ValueError("Лист 'Аксессуары Apple' не найден в файле.")

    sheet = workbook["Аксессуары Apple"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()

                price = parse_price(price_raw)
                if price is None:
                    continue

                results.append({
                    "raw_data": raw_text,
                    "model": raw_text,
                    "price": price,
                })

    return results


# ── iMac ────────────────────────────────────────────────


def parse_imac_string(raw_text):
    """
    Разбирает строку с моделью iMac.
    Примеры:
      "iMac M3 8/256GB (8/8) Green MQRA3"
      "iMac M4 24/1TB (10/10) Silver"
      "iMac M2 16/512GB Blue"
      "iMac M3 8/256GB (8/8) Green"
    Возвращает (model, processor, memory, storage, cpu_cores, gpu_cores, color, part_number).
    """
    colors_multi = r'Stack\s+(?:Blue|Green|Purple)|Blue\s+Violet'
    colors_single = r'Silver|Gold|Green|Blue|Pink|Purple|Red|White|Black|Grey|Gray|Yellow|Orange|Midnight|Starlight'
    color_pat = rf'({colors_multi}|{colors_single})'

    sto = r'(\d+/\d+(?:GB|TB))'
    cores = r'\((\d+)/(\d+)\)'
    part = r'([A-Z0-9]{5,})'

    pats = [
        # chip ram/storage cores color part_number
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}\s+{cores}\s+{color_pat}\s+{part}$',
        # chip ram/storage cores color
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}\s+{cores}\s+{color_pat}$',
        # chip ram/storage cores part_number
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}\s+{cores}\s+{part}$',
        # chip ram/storage cores only
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}\s+{cores}$',
        # chip ram/storage color part_number
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}\s+{color_pat}\s+{part}$',
        # chip ram/storage color
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}\s+{color_pat}$',
        # chip ram/storage part_number
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}\s+{part}$',
        # chip ram/storage only
        rf'^iMac\s+(M\d+(?:\s+(?:Pro|Max|Ultra))?)\s+{sto}$',
    ]

    for i, pat in enumerate(pats):
        m = re.search(pat, raw_text, re.IGNORECASE)
        if not m:
            continue

        chip = m.group(1).strip()
        ram_storage = m.group(2).strip()

        if "/" in ram_storage:
            parts = ram_storage.split("/", 1)
            memory = parts[0] + "GB"
            storage = parts[1]
        else:
            memory = ram_storage
            storage = "Не указано"

        if i == 0:
            return ("iMac", chip, memory, storage,
                    m.group(3).strip(), m.group(4).strip(),
                    m.group(5).strip(), m.group(6).strip())
        if i == 1:
            return ("iMac", chip, memory, storage,
                    m.group(3).strip(), m.group(4).strip(),
                    m.group(5).strip(), "Не указано")
        if i == 2:
            return ("iMac", chip, memory, storage,
                    m.group(3).strip(), m.group(4).strip(),
                    "Не указано", m.group(5).strip())
        if i == 3:
            return ("iMac", chip, memory, storage,
                    m.group(3).strip(), m.group(4).strip(),
                    "Не указано", "Не указано")
        if i == 4:
            return ("iMac", chip, memory, storage,
                    "Не указано", "Не указано",
                    m.group(3).strip(), m.group(4).strip())
        if i == 5:
            return ("iMac", chip, memory, storage,
                    "Не указано", "Не указано",
                    m.group(3).strip(), "Не указано")
        if i == 6:
            return ("iMac", chip, memory, storage,
                    "Не указано", "Не указано",
                    "Не указано", m.group(3).strip())
        if i == 7:
            return ("iMac", chip, memory, storage,
                    "Не указано", "Не указано",
                    "Не указано", "Не указано")

    return ("iMac", "Не указано", "Не указано", "Не указано",
            "Не указано", "Не указано", "Не указано", "Не указано")


def parse_imac_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "iMac" not in workbook.sheetnames:
        raise ValueError("Лист 'iMac' не найден в файле.")

    sheet = workbook["iMac"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model, processor, memory, storage, cpu_cores, gpu_cores, color, part_number = parse_imac_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                country = resolve_country(country_raw)

                results.append({
                    "raw_data": f"{raw_text} | {country}",
                    "model": model,
                    "processor": processor,
                    "memory": memory,
                    "storage": storage,
                    "cpu_cores": cpu_cores,
                    "gpu_cores": gpu_cores,
                    "color": color,
                    "part_number": part_number,
                    "price": price,
                    "country": country,
                })

    return results


# ── Apple Watch ─────────────────────────────────────────

def parse_watch_string(raw_text):
    size = r"(\d{2,3})mm"
    band_sizes = r"(XS|S|M|L|XL|S/M|M/L|S-L|M-L|Long|Short|Standard|One\s+Size|Adjustable)"
    case_colors = r"(Jet\s+Black|Rose\s+Gold|Silver|Gold|Space\s+Black|Space\s+Gray|Midnight|Starlight|Product\s+RED|Deep\s+Purple|Sierra\s+Blue|Coral|Yellow|Red|Navy|Cream|Tan|Alpine\s+Green|White|Blue|Green|Pink|Purple|Black|Graphite|Graphite|Space|Natural\s+Ti|Titanium|Orange|Nectarine|Charcoal|Fog|Dark\s+Night)"
    band_types = r"(Solo\s+Loop|Braided\s+Solo\s+Loop|Sport\s+Loop|Sport\s+Band|Alpine\s+Loop|Trail\s+Loop|Ocean\s+Band|Milanese\s+Loop|Link\s+Bracelet|Woven\s+Loop|Modern\s+Buckle|Strap|Loop|Band|Bracelet)"

    pats = [
        # Ultra 2 49mm Natural Ti Alpine Loop Medium
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+Natural\s+Ti\s+(?:Tan|Green|Orange|White|Dark\s+Night)\s+(Alpine|Trail|Sport|Ocean)\s+(Loop|Band)\s+{band_sizes}$",
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+Natural\s+Ti\s+(?:Tan|Green|Orange|White|Dark\s+Night)\s+(Alpine|Trail|Sport|Ocean)\s+(Loop|Band)$",
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+Natural\s+Ti\s+(Alpine|Trail|Sport|Ocean)\s+(Loop|Band)\s+{band_sizes}$",
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+Natural\s+Ti\s+(Alpine|Trail|Sport|Ocean)\s+(Loop|Band)$",
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+{case_colors}\s+(Alpine|Trail|Sport|Ocean)\s+(Loop|Band)\s+{band_sizes}$",
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+{case_colors}\s+(Alpine|Trail|Sport|Ocean)\s+(Loop|Band)$",
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+{case_colors}\s+(.+?)$",
        rf"^(Ultra)\s*(\d+)?\s+{size}\s+(.+?)$",
        # SE2/SE3: "SE2 2023 40mm Starlight Sport Band (S/M)"
        rf"^SE(\d+)\s+(\d{{4}})\s+{size}\s+{case_colors}\s+{band_types}(?:\s+\(?{band_sizes}\)?)?$",
        rf"^SE(\d+)\s+(\d{{4}})\s+{size}\s+{case_colors}\s+{band_types}$",
        rf"^SE(\d+)\s+(\d{{4}})\s+{size}\s+{case_colors}\s+\(?{band_sizes}\)?$",
        rf"^SE(\d+)\s+(\d{{4}})\s+{size}\s+{case_colors}$",
        rf"^SE(\d+)\s+(\d{{4}})\s+{size}\s+(.*?)$",
        # S10 42mm Jet Black Sport Loop M/L
        rf"^S(\d+|E)\s+{size}\s+{case_colors}\s+{band_types}\s+{band_sizes}$",
        rf"^S(\d+|E)\s+{size}\s+{case_colors}\s+{band_types}$",
        rf"^S(\d+|E)\s+{size}\s+{case_colors}\s+(.*?)$",
        rf"^S(\d+|E)\s+{size}\s+{case_colors}$",
        rf"^S(\d+|E)\s+{size}\s+(.*?)$",
        # SE 40mm Silver White Sport Band M/L
        rf"^SE\s+{size}\s+{case_colors}\s+{band_types}\s+{band_sizes}$",
        rf"^SE\s+{size}\s+{case_colors}\s+{band_types}$",
        rf"^SE\s+{size}\s+{case_colors}\s+(.*?)$",
        rf"^SE\s+{size}\s+{case_colors}$",
        rf"^SE\s+{size}\s+(.*?)$",
    ]

    ultrawords = {"alpine", "trail", "ocean", "sport", "loop", "band", "bracelet", "milanese", "woven", "modern", "buckle", "strap", "solo"}
    sizewords = {"xs", "s", "m", "l", "xl", "s/m", "m/l", "s-l", "m-l", "long", "short", "standard", "one size", "adjustable"}

    for pat in pats:
        m = re.search(pat, raw_text, re.IGNORECASE)
        if m:
            return m.groups()

    return (raw_text,)


def parse_watch_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "Apple Watch" not in workbook.sheetnames:
        raise ValueError("Лист 'Apple Watch' не найден в файле.")

    sheet = workbook["Apple Watch"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                groups = parse_watch_string(raw_text)
                if not groups:
                    continue

                price = parse_price(price_raw)
                if price is None:
                    continue

                country = resolve_country(country_raw)

                ng = len(groups)
                entry = {"raw_data": f"{raw_text} | {country}", "price": price, "country": country}

                if ng >= 1 and groups[0]:
                    series = groups[0]
                    if series == "Ultra":
                        ultra_gen = ""
                        size_val = ""
                        color_val = ""
                        band = ""

                        if ng >= 2 and groups[1]:
                            ultra_gen = groups[1].strip()
                        entry["series"] = "Ultra"
                        if ultra_gen:
                            entry["generation"] = ultra_gen

                        # Ultra patterns: g = (Ultra, gen?, size, color?, band_name?, loop/band?, band_size)?
                        idx = 2  # after Ultra and gen
                        if idx < ng and groups[idx]:
                            entry["size"] = str(groups[idx]).strip()
                            idx += 1
                        if idx < ng and groups[idx]:
                            color_val = groups[idx].strip()
                            # check if this is actually a band word, not color
                            if color_val.lower() in {"alpine", "trail", "ocean", "sport", "loop", "band"}:
                                band = color_val
                                color_val = ""
                            else:
                                entry["color"] = color_val
                                entry["case_material"] = "Titanium"
                                idx += 1
                        if idx < ng and groups[idx]:
                            bt = groups[idx].strip()
                            if bt.lower() in {"loop", "band"}:
                                bname = groups[idx - 1].strip() if idx > 0 else ""
                                band = f"{bname} {bt}".strip() if bname else bt
                            else:
                                band = bt
                            if band:
                                entry["band_info"] = band
                            idx += 1
                        if idx < ng and groups[idx]:
                            bs = groups[idx].strip()
                            if band:
                                entry["band_info"] = f"{band} ({bs})"

                        model_name = f"Apple Watch Ultra {ultra_gen}".strip() if ultra_gen else "Apple Watch Ultra"
                        entry["model"] = model_name
                    elif re.match(r"^SE\d+\b", raw_text, re.IGNORECASE):
                        # SE2/SE3: "SE2 2023 40mm Starlight Sport Band (S/M)"
                        # g = (num, year, size, color, [band_type/band_size], [band_size])
                        series_clean = "SE" + str(series).strip()
                        model_name = f"Apple Watch {series_clean}"
                        entry["series"] = series_clean

                        if ng >= 2 and groups[1]:
                            entry["year"] = str(groups[1]).strip()
                        if ng >= 3 and groups[2]:
                            entry["size"] = str(groups[2]).strip()
                        if ng >= 4 and groups[3]:
                            entry["color"] = groups[3].strip()
                        band = ""
                        band_size = ""
                        if ng >= 5 and groups[4]:
                            band = groups[4].strip()
                        if ng >= 6 and groups[5]:
                            band_size = groups[5].strip()
                        if band_size:
                            if band:
                                entry["band_info"] = f"{band} ({band_size})"
                            else:
                                entry["band_info"] = f"({band_size})"
                        elif band:
                            # Паттерн без типа ремешка: остаток — размер (S/M), тип
                            # ремешка или просто текст. Размеры содержат "/".
                            if re.match(r"^[A-Za-z]+/[A-Za-z]+$", band):
                                entry["band_info"] = f"({band})"
                            else:
                                entry["band_info"] = band

                        entry["model"] = model_name
                    else:
                        # S or SE
                        if series == "E":
                            model_name = "Apple Watch SE"
                        else:
                            model_name = f"Apple Watch S{series}"
                        entry["series"] = series

                        if ng >= 2 and groups[1]:
                            entry["size"] = str(groups[1]).strip()
                        if ng >= 3 and groups[2]:
                            entry["color"] = groups[2].strip()
                        band_parts = []
                        if ng >= 4 and groups[3]:
                            band_parts.append(groups[3].strip())
                        if ng >= 5 and groups[4]:
                            band_parts.append(groups[4].strip())
                        if ng >= 6 and groups[5]:
                            band_parts.append(groups[5].strip())
                        if band_parts:
                            entry["band_info"] = " ".join(band_parts)

                        entry["model"] = model_name

                results.append(entry)

    return results


# ── AirPods ──────────────────────────────────────────────

AIRPODS_COLORS = {
    "Midnight": "Midnight",
    "Starlight": "Starlight",
    "White": "White",
    "Silver": "Silver",
    "Space Gray": "Space Gray",
    "Sky Blue": "Sky Blue",
    "Pink": "Pink",
    "Green": "Green",
    "Orange": "Orange",
    "Product RED": "Product RED",
    "Purple": "Purple",
    "Blue": "Blue",
    "Lightning": "Lightning",
    "Magenta": "Magenta",
    "Yellow": "Yellow",
    "Cyan": "Cyan",
    "Rose Gold": "Rose Gold",
    "Gold": "Gold",
    "Black": "Black",
    "Red": "Red",
    "Coral": "Coral",
}

AIRPODS_PATTERNS = [
    # AirPods Max 2 Midnight
    rf"^(AirPods\s+Max)\s+(\d+)\s+({ '|'.join(AIRPODS_COLORS.keys()) })$",
    # AirPods Max 2
    rf"^(AirPods\s+Max)\s+(\d+)$",
    # AirPods Pro 2 USB-C Midnight
    rf"^(AirPods\s+Pro)\s+(\d+)(?:\s+([\w\s-]+?))?\s+({ '|'.join(AIRPODS_COLORS.keys()) })$",
    # AirPods Pro 2 USB-C
    rf"^(AirPods\s+Pro)\s+(\d+)(?:\s+([\w\s-]+?))?\s*$",
    # AirPods 4 ANC Midnight
    rf"^(AirPods)\s+(\d+)(?:\s+([\w\s-]+?))?\s+({ '|'.join(AIRPODS_COLORS.keys()) })$",
    # AirPods 4
    rf"^(AirPods)\s+(\d+)(?:\s+([\w\s-]+?))?\s*$",
    # AirPods Max Midnight
    rf"^(AirPods\s+Max)\s+({ '|'.join(AIRPODS_COLORS.keys()) })$",
    # AirPods Max
    rf"^(AirPods\s+Max)$",
    # AirPods Pro 2 Midnight
    rf"^(AirPods\s+Pro)\s+(\d+)\s+({ '|'.join(AIRPODS_COLORS.keys()) })$",
    # AirPods Pro Midnight
    rf"^(AirPods\s+Pro)\s+({ '|'.join(AIRPODS_COLORS.keys()) })$",
    # AirPods Pro
    rf"^(AirPods\s+Pro)$",
]


def parse_airpods_string(raw_text):
    """
    Разбирает строку с AirPods.
    Возвращает (display_model, color).
    display_model — полное имя без цвета (AirPods Max 2, AirPods Pro 2, AirPods 3, ...)
    color — цвет или None.
    """
    for pat in AIRPODS_PATTERNS:
        m = re.search(pat, raw_text, re.IGNORECASE)
        if m:
            parts = [g for g in m.groups() if g]
            # remove trailing if it's a color
            color = None
            if parts and parts[-1] in AIRPODS_COLORS:
                color = parts.pop()
            display = " ".join(parts)
            return display, color

    return raw_text, None


def parse_airpods_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "AirPods" not in workbook.sheetnames:
        raise ValueError("Лист 'AirPods' не найден в файле.")

    sheet = workbook["AirPods"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                display_model, color = parse_airpods_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                country = resolve_country(country_raw)

                entry = {
                    "raw_data": f"{raw_text} | {country}",
                    "model": display_model,
                    "price": price,
                    "country": country,
                }

                if color:
                    entry["color"] = color

                results.append(entry)

    return results


# ── Samsung ─────────────────────────────────────────────

SAMSUNG_COLORS = [
    "White Silver",
    "Silver Blue",
    "Pink Gold",
    "Jet Black",
    "Icy Blue",
    "Ice Blue",
    "Light Gray",
    "Light Grey",
    "Space Black",
    "Space Gray",
    "Rose Gold",
    "Sky Blue",
    "Deep Purple",
    "Sierra Blue",
    "Alpine Green",
    "Product RED",
    "Midnight",
    "Starlight",
    "Silver",
    "Graphite",
    "Blueblack",
    "Gold",
    "Black",
    "White",
    "Blue",
    "Green",
    "Gray",
    "Grey",
    "Violet",
    "Navy",
    "Lime",
    "Lavender",
    "Mint",
    "Olive",
    "Pink",
    "Yellow",
    "Red",
    "Orange",
    "Coral",
    "Purple",
    "Brown",
    "Cream",
    "Beige",
    "Teal",
    "Cyan",
    "Magenta",
    "Rose",
]


def parse_samsung_string(raw_text):
    """
    Разбирает строку с товаром Samsung.
    Примеры:
      "S25 Ultra 12/1TB White Silver"   → phone
      "A36 8/256GB Black"               → phone
      "Z Fold 7 12/512GB Blue"          → phone
      "Watch 8 Ultra 47mm LTE Blue"     → watch
      "Watch 8 Classic 46mm Black"      → watch
      "Buds 3 Pro Silver"               → buds
      "Buds FE Gray"                    → buds
    Возвращает (model, memory, storage, color, size, connectivity, product_type).
    """
    ram_storage = re.search(r'(\d+)/(\d+(?:GB|TB))', raw_text, re.IGNORECASE)
    if ram_storage:
        model_part = raw_text[:ram_storage.start()].strip()
        memory = f"{ram_storage.group(1)}GB"
        storage = ram_storage.group(2).upper()
        color = raw_text[ram_storage.end():].strip() or "Не указано"
        return model_part, memory, storage, color, "Не указано", "Не указано", "phone"

    size_match = re.search(r'(\d{2,3})mm', raw_text, re.IGNORECASE)
    if size_match:
        model_part = raw_text[:size_match.start()].strip()
        size = f"{size_match.group(1)}mm"
        rest = raw_text[size_match.end():].strip()
        connectivity = "Не указано"
        color = rest or "Не указано"
        conn_match = re.match(r'(LTE|5G|Cellular|Wi-Fi|BT)\s*(.*)$', rest, re.IGNORECASE)
        if conn_match:
            connectivity = conn_match.group(1).upper()
            color = conn_match.group(2).strip() or "Не указано"
        return model_part, "Не указано", "Не указано", color, size, connectivity, "watch"

    for color_name in SAMSUNG_COLORS:
        if raw_text.lower().endswith(color_name.lower()):
            idx = len(raw_text) - len(color_name)
            if idx == 0 or raw_text[idx - 1].isspace():
                model_part = raw_text[:idx].strip()
                return model_part, "Не указано", "Не указано", color_name, "Не указано", "Не указано", "buds"

    return raw_text, "Не указано", "Не указано", "Не указано", "Не указано", "Не указано", "other"


def parse_samsung_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "Samsung" not in workbook.sheetnames:
        raise ValueError("Лист 'Samsung' не найден в файле.")

    sheet = workbook["Samsung"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model_part, memory, storage, color, size, connectivity, product_type = parse_samsung_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                if model_part.lower().startswith("samsung"):
                    full_model = model_part
                else:
                    full_model = f"Samsung {model_part}"

                country = resolve_country(country_raw)

                entry = {
                    "raw_data": f"{raw_text} | {country}",
                    "model": full_model,
                    "price": price,
                    "country": country,
                }

                if memory != "Не указано":
                    entry["memory"] = memory
                if storage != "Не указано":
                    entry["storage"] = storage
                if color != "Не указано":
                    entry["color"] = color
                if size != "Не указано":
                    entry["size"] = size
                if connectivity != "Не указано":
                    entry["connectivity"] = connectivity

                results.append(entry)

    return results


# ── Игровые приставки ──────────────────────────────────


def parse_console_string(raw_text):
    """
    Разбирает строку с игровой приставкой.
    Примеры:
      "PS5 Slim Digital Edition 1TB"
      "PS5 Slim Digital Edition 1TB White"
      "PS5 Pro Digital Edition 2TB"
      "PlayStation VR2"
      "DualSense Charging Station"
      "XBOX Series X 1TB Black"
      "Oculus Quest 3S 128GB"
      "Nintendo Switch Lite Hyrule Edition"
      "Nintendo Switch 2 Mario Kart World Bundle"
      "Steam Deck (OLED) 512GB"
    Возвращает (model, storage, color, edition).
    """
    m = re.match(r'^(PS5\s+(?:Slim|Pro))\s+(Digital\s+Edition|Disc\s+Edition)\s+(\d+(?:GB|TB))(?:\s+(White|Black))?$', raw_text, re.IGNORECASE)
    if m:
        return m.group(1).strip(), m.group(3).upper(), m.group(4) or "Не указано", m.group(2)

    m = re.match(r'^(PlayStation\s+VR2)$', raw_text, re.IGNORECASE)
    if m:
        return m.group(1), "Не указано", "Не указано", "Не указано"

    m = re.match(r'^(XBOX\s+Series\s+X)\s+(\d+(?:GB|TB))(?:\s+(White|Black))?$', raw_text, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2).upper(), m.group(3) or "Не указано", "Не указано"

    m = re.match(r'^(Oculus\s+Quest\s+\d+\S*)\s+(\d+(?:GB|TB))$', raw_text, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2).upper(), "Не указано", "Не указано"

    m = re.match(r'^(Steam\s+Deck\s+\(OLED\))(?:\s+(\d+(?:GB|TB)))?$', raw_text, re.IGNORECASE)
    if m:
        return m.group(1), m.group(2).upper() if m.group(2) else "Не указано", "Не указано", "Не указано"

    m = re.match(r'^(DualSense)(?:\s+(Charging\s+Station))?(?:\s+(White|Black))?$', raw_text, re.IGNORECASE)
    if m:
        return m.group(1), "Не указано", m.group(3) or "Не указано", m.group(2) or "Не указано"

    m = re.match(r'^(Nintendo\s+Switch(?:\s+Lite|\s+2)?)(?:\s+(.+))?$', raw_text, re.IGNORECASE)
    if m:
        extra = m.group(2).strip() if m.group(2) else "Не указано"
        return m.group(1), "Не указано", "Не указано", extra

    return raw_text, "Не указано", "Не указано", "Не указано"


def parse_consoles_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "Игровые приставки" not in workbook.sheetnames:
        raise ValueError("Лист 'Игровые приставки' не найден в файле.")

    sheet = workbook["Игровые приставки"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model, storage, color, edition = parse_console_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                entry = {
                    "raw_data": raw_text,
                    "model": model,
                    "price": price,
                }

                if storage != "Не указано":
                    entry["storage"] = storage
                if color != "Не указано":
                    entry["color"] = color
                if edition != "Не указано":
                    entry["edition"] = edition

                country = resolve_country(country_raw)
                if country != "Не указана":
                    entry["country"] = country
                    entry["raw_data"] = f"{raw_text} | {country}"

                results.append(entry)

    return results


# ── Dyson ───────────────────────────────────────────────


def parse_dyson_string(raw_text):
    """
    Разбирает строку с товаром Dyson.
    Примеры:
      "Airstrait HT01 Nickel/Cooper"
      "Supersonic HD07 Nickel/Cooper"
      "Supersonic HD07 Blue/Copper (с кейсом)"
      "Supersonic HD16 Pink/Rose Gold"
      "Airwrap HS05 Long Nickel/Copper"
      "Airwrap HS08 Long Blue/Copper diffuser"
      "V8 Advanced SV25 Silver/Nickel"
      "V12 Detect Slim Absolute SV46 Yellow/Nickel"
      "PH05 White/Gold"
      "PH05"
      "HU02 White"
    Возвращает (model, color, accessory).
    """
    accessory = "Не указано"
    text = raw_text.strip()
    for suffix in ["(с кейсом)", "diffuser"]:
        if text.lower().endswith(suffix):
            accessory = suffix
            text = text[: -len(suffix)].strip()
            break

    code_match = re.search(r'[A-Z]{2}\d{2}', text)
    if code_match:
        model = text[:code_match.end()].strip()
        rest = text[code_match.end():].strip()
        if rest.lower().startswith("long"):
            model = f"{model} Long"
            rest = rest[4:].strip()
        color = rest or "Не указано"
        return model, color, accessory

    return raw_text, "Не указано", accessory


def parse_dyson_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "Dyson" not in workbook.sheetnames:
        raise ValueError("Лист 'Dyson' не найден в файле.")

    sheet = workbook["Dyson"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model_part, color, accessory = parse_dyson_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                if model_part.lower().startswith("dyson"):
                    full_model = model_part
                else:
                    full_model = f"Dyson {model_part}"

                country = resolve_country(country_raw)

                entry = {
                    "raw_data": f"{raw_text} | {country}",
                    "model": full_model,
                    "price": price,
                    "country": country,
                }

                if color != "Не указано":
                    entry["color"] = color
                if accessory != "Не указано":
                    entry["accessory"] = accessory

                results.append(entry)

    return results


# ── Xiaomi ──────────────────────────────────────────────


def parse_xiaomi_string(raw_text):
    """
    Разбирает строку с товаром Xiaomi/Redmi.
    Примеры:
      "Redmi Pad Pro 8/256GB Green"
      "Redmi Note 14 Pro Plus 5G 12/512GB Purple"
      "Xiaomi Pad 7 8/256GB Gray Wi-Fi"
      "Xiaomi 13 8/256GB Black"
    Возвращает (model, memory, storage, color, connectivity).
    """
    ram_storage = re.search(r'(\d+)/(\d+(?:GB|TB))', raw_text, re.IGNORECASE)
    if not ram_storage:
        return raw_text, "Не указано", "Не указано", "Не указано", "Не указано"

    model_part = raw_text[:ram_storage.start()].strip()
    memory = f"{ram_storage.group(1)}GB"
    storage = ram_storage.group(2).upper()
    rest = raw_text[ram_storage.end():].strip()

    connectivity = "Не указано"
    color = rest
    conn_match = re.search(r'\s+(Wi-Fi|WiFi|5G|LTE|Cellular)\s*$', rest, re.IGNORECASE)
    if conn_match:
        connectivity = conn_match.group(1)
        color = rest[:conn_match.start()].strip()

    return model_part, memory, storage, color or "Не указано", connectivity


def parse_xiaomi_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "Xiaomi" not in workbook.sheetnames:
        raise ValueError("Лист 'Xiaomi' не найден в файле.")

    sheet = workbook["Xiaomi"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model_part, memory, storage, color, connectivity = parse_xiaomi_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                if model_part.lower().startswith(("xiaomi", "redmi")):
                    full_model = model_part
                else:
                    full_model = f"Xiaomi {model_part}"

                country = resolve_country(country_raw)

                entry = {
                    "raw_data": f"{raw_text} | {country}",
                    "model": full_model,
                    "price": price,
                    "country": country,
                }

                if memory != "Не указано":
                    entry["memory"] = memory
                if storage != "Не указано":
                    entry["storage"] = storage
                if color != "Не указано":
                    entry["color"] = color
                if connectivity != "Не указано":
                    entry["connectivity"] = connectivity

                results.append(entry)

    return results


# ── POCO ────────────────────────────────────────────────


def parse_poco_string(raw_text):
    """
    Разбирает строку с товаром POCO.
    Примеры:
      "Poco C85 6/128GB Green"
      "Poco X7 Pro 8/256GB Green"
      "Poco F7 Ultra 12/256GB Black"
      "POCO Pad 8/256GB Gray"
    Возвращает (model, memory, storage, color, connectivity).
    """
    ram_storage = re.search(r'(\d+)/(\d+(?:GB|TB))', raw_text, re.IGNORECASE)
    if not ram_storage:
        return raw_text, "Не указано", "Не указано", "Не указано", "Не указано"

    model_part = raw_text[:ram_storage.start()].strip()
    memory = f"{ram_storage.group(1)}GB"
    storage = ram_storage.group(2).upper()
    rest = raw_text[ram_storage.end():].strip()

    connectivity = "Не указано"
    color = rest
    conn_match = re.search(r'\s+(Wi-Fi|WiFi|5G|LTE|Cellular)\s*$', rest, re.IGNORECASE)
    if conn_match:
        connectivity = conn_match.group(1)
        color = rest[:conn_match.start()].strip()

    return model_part, memory, storage, color or "Не указано", connectivity


def parse_poco_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "POCO" not in workbook.sheetnames:
        raise ValueError("Лист 'POCO' не найден в файле.")

    sheet = workbook["POCO"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model_part, memory, storage, color, connectivity = parse_poco_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                if model_part.lower().startswith("poco"):
                    full_model = "POCO" + model_part[4:]
                else:
                    full_model = model_part

                country = resolve_country(country_raw)

                entry = {
                    "raw_data": f"{raw_text} | {country}",
                    "model": full_model,
                    "price": price,
                    "country": country,
                }

                if memory != "Не указано":
                    entry["memory"] = memory
                if storage != "Не указано":
                    entry["storage"] = storage
                if color != "Не указано":
                    entry["color"] = color
                if connectivity != "Не указано":
                    entry["connectivity"] = connectivity

                results.append(entry)

    return results


# ── Яндекс Станции ─────────────────────────────────────


def parse_station_string(raw_text):
    """
    Разбирает строку с Яндекс Станцией.
    Примеры:
      "Яндекс Станция Лайт 2 Без Часов (Зелёный)"
      "Яндекс Станция Лайт 2 (Синий)"
      "Яндекс Станция Мини 2 (Чёрный)"
      "Яндекс Станция Мини 3 + Часы (Чёрный)"
      "Яндекс Станция Миди (Малиновый)"
    Возвращает (model, color, clock).
    """
    color_match = re.search(r'\(([^)]+)\)\s*$', raw_text)
    color = color_match.group(1).strip() if color_match else "Не указано"
    body = raw_text[:color_match.start()].strip() if color_match else raw_text.strip()

    clock = "Не указано"
    clock_match = re.search(r'Без\s+Часов', body, re.IGNORECASE)
    if clock_match:
        clock = "Без Часов"
        body = body[:clock_match.start()].strip()
    else:
        clock_match = re.search(r'\+\s*Часы', body, re.IGNORECASE)
        if clock_match:
            clock = "С Часами"
            body = body[:clock_match.start()].strip()

    return body.strip(), color, clock


def parse_stations_excel_to_dicts(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл '{file_path}' не найден.")

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    if "Яндекс Станции" not in workbook.sheetnames:
        raise ValueError("Лист 'Яндекс Станции' не найден в файле.")

    sheet = workbook["Яндекс Станции"]
    results = []

    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        model_raw = row[0]
        price_raw = row[1]
        country_raw = row[2]

        if price_raw is not None and str(price_raw).strip() != "":
            if model_raw:
                raw_text = str(model_raw).strip()
                model, color, clock = parse_station_string(raw_text)

                price = parse_price(price_raw)
                if price is None:
                    continue

                country = resolve_country(country_raw)

                entry = {
                    "raw_data": f"{raw_text} | {country}",
                    "model": model,
                    "price": price,
                    "country": country,
                }

                if color != "Не указано":
                    entry["color"] = color
                if clock != "Не указано":
                    entry["clock"] = clock

                results.append(entry)

    return results
